"""
Interactive prize draw: compact chroma overlay, HTML wheel for OBS, prize board window, and Spin & controls window.
"""

from __future__ import annotations

import mimetypes
import os
import random
import sys
import re
import ssl
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Literal
from zoneinfo import ZoneInfo
import urllib.request
import tkinter as tk
import tkinter.font as tkfont
import webbrowser
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk
from urllib.parse import quote, urlparse, unquote
from urllib.request import url2pathname

import app_local_state
import draw_prize

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]


WHEEL_FRAME_BG = "#1a1a2e"
# Solid dark green behind wheel chrome for OBS Chroma Key. Avoid this exact color in prize art.
OBS_CHROMA_KEY = "#196619"
WHEEL_BG = "#16213e"
WHEEL_CELL_BG = "#0f3460"
WHEEL_CELL_BORDER = "#533483"
WHEEL_FG = "#eaeaea"
WHEEL_WIN_BG = "#e94560"
WHEEL_WIN_FG = "#ffffff"
WHEEL_POINTER = "#ffd93d"
WHEEL_ACCENT = "#ff6b6b"
WHEEL_TITLE = "#ffd93d"
WHEEL_MUTED = "#4a4e69"

# Horizontal prize strip (canvas): slate / indigo — used only for wheel drawing (chroma canvas bg stays OBS_CHROMA_KEY).
WHEEL_STRIP_SLOT_BG = "#1a2744"
WHEEL_STRIP_SLOT_BD = "#334155"
WHEEL_STRIP_FG = "#e2e8f0"
WHEEL_STRIP_WIN_BG = "#4f46e5"
WHEEL_STRIP_WIN_FG = "#f8fafc"
WHEEL_STRIP_WIN_RING = "#a5b4fc"
WHEEL_STRIP_POINTER = "#fbbf24"
WHEEL_STRIP_POINTER_EDGE = "#0f172a"
WHEEL_STRIP_TRACK = "#273549"
# Wheel strip vertical scale (cell height + pads + overlay window vs original design).
WHEEL_VERTICAL_SCALE = 1.2

WHEEL_STRIP_PAD_TOP = int(round(28 * WHEEL_VERTICAL_SCALE))
WHEEL_STRIP_PAD_BOTTOM = int(round(14 * WHEEL_VERTICAL_SCALE))
WHEEL_STRIP_CELL_GAP = 6.0
# HTML / OBS Browser Source: strip only — base scale vs. Tk, then width/pad tweaks for overlay layout.
HTML_WHEEL_DISPLAY_SCALE = 1.25
HTML_WHEEL_STRIP_WIDTH_MUL = 0.8  # strip cells 20% narrower than (Tk × DISPLAY_SCALE)
HTML_WHEEL_STRIP_PAD_V_MUL = 1.1  # vertical strip padding +10% vs. (Tk pad × DISPLAY_SCALE)
# HTML strip row height: divide snapshot pad top/bottom by this so cells are ~30% taller (same scroll math).
HTML_WHEEL_STRIP_CELL_HEIGHT_MUL = 1.3
BTN_SPIN = "#2980b9"
BTN_SUPER = "#8e44ad"
BTN_REROLL = "#16a085"
BTN_KEEP = "#c0392b"
BTN_SKIP = "#dc7633"
BTN_UNDO = "#1f618d"
BTN_DISABLED = "#3d3d54"
DRAG_BAR = "#f39c12"

# Prize board storage grid (row-major: `cols` tiles per row, then wrap).
INV_GRID_GAP = 2
INV_SLOT_INSET = 2
INV_SLOT_EDGE = "#7a9fd6"
INV_SLOT_FACE = "#0c1220"
INV_QTY_BADGE_BG = "#ffd93d"
INV_QTY_BADGE_FG = "#1a1a2e"

# Prize board grid: columns = items per row; rows = target row count for sizing tiles to window height.
PRIZE_BOARD_SLOT_MIN_PX = 44
PRIZE_BOARD_SLOT_MAX_PX = 160
PRIZE_BOARD_GRID_COLS_DEFAULT = 6
PRIZE_BOARD_ROWS_FIT_DEFAULT = 4

# Prize board Toplevel: readable on a monitor (not keyed for OBS chroma).
PRIZE_BOARD_WINDOW_BG = WHEEL_FRAME_BG
PRIZE_BOARD_CONTENT_BG = WHEEL_BG
PRIZE_BOARD_HEADER_BG = "#152238"
PRIZE_BOARD_RIM = WHEEL_CELL_BORDER

# Spin & controls Toplevel: same on-monitor palette; cards sit slightly inset.
SPIN_CONTROLS_CARD_BG = "#131c2e"
SPIN_CONTROLS_CARD_BORDER = WHEEL_CELL_BORDER
# Default / minimum size — window is user-resizable (drag edges or corners to expand).
SPIN_CONTROLS_WIN_WIDTH = 720
SPIN_CONTROLS_WIN_HEIGHT = 920
SPIN_CONTROLS_WIN_MIN_WIDTH = 560
SPIN_CONTROLS_WIN_MIN_HEIGHT = 580
# Live wheel on Spin & controls only (smaller than the hidden OBS/snapshot canvas).
SPIN_CONTROLS_WHEEL_CANVAS_H = int(round(128 * WHEEL_VERTICAL_SCALE))
SPIN_CONTROLS_WHEEL_CANVAS_H_MAX = int(round(168 * WHEEL_VERTICAL_SCALE))
# Overlay window: compact height vs setup panel scroll viewport (Step 1, 2, options, log).
OVERLAY_WIDTH = 480
OVERLAY_HEIGHT_COMPACT = int(round(280 * WHEEL_VERTICAL_SCALE))
OVERLAY_SETUP_SCROLL_MAX = 360


def script_dir() -> Path:
    """Folder for user data (winner_sessions, default browse dir): next to the .exe when frozen."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_root() -> Path:
    """Read-only bundled assets (e.g. web/); PyInstaller one-file extract dir when frozen."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS).resolve()
    return Path(__file__).resolve().parent


def _pillow_runtime_available() -> bool:
    try:
        import PIL.Image  # noqa: F401
        import PIL.ImageTk  # noqa: F401

        return True
    except ImportError:
        return False


def _looks_like_html_head(data: bytes) -> bool:
    head = data[:800].lstrip().lower()
    return head.startswith((b"<", b"<!doctype")) or b"<html" in head[:1200]


def _normalize_image_ref(ref: str) -> str:
    ref = (ref or "").strip().strip("\ufeff").strip('"').strip("'")
    if not ref:
        return ""
    if ref.lower().startswith("file:"):
        parsed = urlparse(ref.replace("\\", "/"))
        raw_path = unquote(parsed.path or "")
        if parsed.netloc and os.name == "nt":
            ref = f"\\\\{parsed.netloc}{raw_path.replace('/', os.sep)}"
        else:
            try:
                ref = url2pathname(raw_path)
            except (ValueError, OSError):
                pass
        return ref.strip()
    m = re.search(r"https?://[^\s]+", ref, flags=re.IGNORECASE)
    if m:
        return m.group(0).rstrip(").,]>'\"")
    return ref


def _path_for_project_join(ref: str) -> str:
    ref = ref.strip()
    if not ref:
        return ref
    if ref.startswith("\\\\") or (len(ref) >= 2 and ref[1] == ":"):
        return ref
    return ref.lstrip("/\\")


class PrizeListOverlay(tk.Toplevel):
    """Separate window: PC-style remaining prizes grid, tuned for viewing on a monitor (not chroma/OBS)."""

    def __init__(self, app: "DrawPrizeApp") -> None:
        super().__init__(app)
        self._app = app
        self.title("Energy Break — Prize board")
        self.minsize(560, 480)
        self.geometry("760x680")
        self.configure(bg=PRIZE_BOARD_WINDOW_BG)
        try:
            self.attributes("-topmost", False)
        except tk.TclError:
            pass
        try:
            self.resizable(True, True)
        except tk.TclError:
            pass

        app.update_idletasks()
        px = app.winfo_rootx() + app.winfo_width() + 16
        py = app.winfo_rooty()
        self.geometry(f"760x680+{px}+{py}")

        self.protocol("WM_DELETE_WINDOW", self._close_overlay)

        outer = tk.Frame(
            self,
            bg=PRIZE_BOARD_WINDOW_BG,
            highlightbackground=PRIZE_BOARD_RIM,
            highlightthickness=1,
        )
        self._void_outer_frame = outer
        outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        inner = tk.Frame(outer, bg=PRIZE_BOARD_CONTENT_BG)
        inner.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        head = tk.Frame(inner, bg=PRIZE_BOARD_HEADER_BG, padx=14, pady=12)
        head.pack(fill=tk.X)
        head_top = tk.Frame(head, bg=PRIZE_BOARD_HEADER_BG)
        head_top.pack(fill=tk.X)
        tk.Label(
            head_top,
            text="Still in draw",
            font=("Segoe UI", 16, "bold"),
            bg=PRIZE_BOARD_HEADER_BG,
            fg=WHEEL_TITLE,
            anchor=tk.W,
        ).pack(side=tk.LEFT)
        app._prize_board_totals_label = tk.Label(
            head_top,
            text="—",
            font=("Segoe UI", 12, "bold"),
            bg=PRIZE_BOARD_HEADER_BG,
            fg=WHEEL_POINTER,
            anchor=tk.E,
        )
        app._prize_board_totals_label.pack(side=tk.RIGHT)
        tk.Label(
            head,
            text="Prizes from your list with quantity remaining (each tile shows stock on hand).",
            font=("Segoe UI", 9),
            bg=PRIZE_BOARD_HEADER_BG,
            fg=WHEEL_MUTED,
            anchor=tk.W,
            justify=tk.LEFT,
        ).pack(fill=tk.X, pady=(8, 0))

        scroll_wrap = tk.Frame(inner, bg=PRIZE_BOARD_CONTENT_BG)
        scroll_wrap.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 14))
        app.remaining_scroll = None  # type: ignore[assignment]
        app.remaining_canvas = None  # type: ignore[assignment]
        app._remaining_canvas_win = None  # type: ignore[assignment]
        app.remaining_inner = tk.Frame(scroll_wrap, bg=PRIZE_BOARD_CONTENT_BG)
        app.remaining_inner.pack(fill=tk.BOTH, expand=True)
        app.remaining_inner.bind("<Configure>", lambda _e: app._schedule_prize_board_layout_refresh())

    def _close_overlay(self) -> None:
        app = self._app
        app._cancel_prize_board_resize_refresh()
        app._prize_win = None
        app._prize_board_totals_label = None  # type: ignore[assignment]
        app.remaining_scroll = None  # type: ignore[assignment]
        app.remaining_canvas = None  # type: ignore[assignment]
        app.remaining_inner = None  # type: ignore[assignment]
        app._remaining_canvas_win = None  # type: ignore[assignment]
        try:
            self.destroy()
        except tk.TclError:
            pass


@dataclass
class _UndoSpinSnapshot:
    """Single-step undo: spin, super keep, skip, or fill-skip (update empty winner row)."""

    kind: Literal["spin", "skip_log", "skip_counter_only", "backfill"]
    list_path: Path
    result: draw_prize.SpinResult | None
    spot_written: int
    winner_row_added: bool


class SpinControlsWindow(tk.Toplevel):
    """Draw / Super / Reroll / Keep, edit spin, fill skipped, HTML wheel, session, prize board — resizable, scrollable."""

    def __init__(self, app: "DrawPrizeApp") -> None:
        super().__init__(app)
        self._app = app
        self.title("Energy Break — Spin & controls")
        win_bg = PRIZE_BOARD_WINDOW_BG
        self.configure(bg=win_bg)
        try:
            self.attributes("-topmost", False)
        except tk.TclError:
            pass
        try:
            self.resizable(True, True)
        except tk.TclError:
            pass
        self.minsize(SPIN_CONTROLS_WIN_MIN_WIDTH, SPIN_CONTROLS_WIN_MIN_HEIGHT)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        app.update_idletasks()
        w, h = SPIN_CONTROLS_WIN_WIDTH, SPIN_CONTROLS_WIN_HEIGHT
        px = max(0, app.winfo_rootx() + 24)
        py = max(0, app.winfo_rooty() + app.winfo_height() + 16)
        self.geometry(f"{w}x{h}+{px}+{py}")

        outer = tk.Frame(
            self,
            bg=win_bg,
            highlightbackground=SPIN_CONTROLS_CARD_BORDER,
            highlightthickness=1,
        )
        outer.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        shell = tk.Frame(outer, bg=PRIZE_BOARD_CONTENT_BG)
        shell.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        header = tk.Frame(shell, bg=PRIZE_BOARD_HEADER_BG)
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="Spin & controls",
            font=("Segoe UI", 17, "bold"),
            bg=PRIZE_BOARD_HEADER_BG,
            fg=WHEEL_TITLE,
            anchor=tk.W,
        ).pack(fill=tk.X, padx=18, pady=(14, 4))
        tk.Label(
            header,
            text="Live wheel preview at the top; draw, edit spin, OBS HTML, new wheel, and prize board below (scroll).",
            font=("Segoe UI", 9),
            bg=PRIZE_BOARD_HEADER_BG,
            fg=WHEEL_MUTED,
            anchor=tk.W,
            justify=tk.LEFT,
        ).pack(fill=tk.X, padx=18, pady=(0, 14))

        scroll_outer = tk.Frame(shell, bg=PRIZE_BOARD_CONTENT_BG)
        scroll_outer.pack(fill=tk.BOTH, expand=True)
        scroll_canvas = tk.Canvas(
            scroll_outer,
            bg=PRIZE_BOARD_CONTENT_BG,
            highlightthickness=0,
            bd=0,
        )
        vscroll = tk.Scrollbar(
            scroll_outer,
            orient=tk.VERTICAL,
            command=scroll_canvas.yview,
            bg=PRIZE_BOARD_CONTENT_BG,
            troughcolor=SPIN_CONTROLS_CARD_BG,
            activebackground=WHEEL_CELL_BG,
        )
        scroll_canvas.configure(yscrollcommand=vscroll.set)
        try:
            scroll_canvas.configure(takefocus=True)
        except tk.TclError:
            pass
        vscroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 2))
        scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._spin_controls_scroll_canvas = scroll_canvas

        body = tk.Frame(scroll_canvas, bg=PRIZE_BOARD_CONTENT_BG)
        body_win = scroll_canvas.create_window((0, 0), window=body, anchor=tk.NW)

        def _sync_scroll_body_width(_event: object | None = None) -> None:
            try:
                cw = max(120, int(scroll_canvas.winfo_width()) - 2)
            except tk.TclError:
                return
            try:
                scroll_canvas.itemconfigure(body_win, width=cw)
            except tk.TclError:
                return

        def _on_scroll_body_configure(_event: tk.Event) -> None:
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))
            _sync_scroll_body_width()

        def _on_scroll_canvas_configure(_event: tk.Event) -> None:
            _sync_scroll_body_width()

        body.bind("<Configure>", _on_scroll_body_configure, add="+")
        scroll_canvas.bind("<Configure>", _on_scroll_canvas_configure, add="+")

        def _wheel_spin_controls(event: tk.Event) -> None:
            if getattr(event, "delta", 0):
                scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _wheel_spin_controls_up(_event: tk.Event) -> None:
            scroll_canvas.yview_scroll(-1, "units")

        def _wheel_spin_controls_down(_event: tk.Event) -> None:
            scroll_canvas.yview_scroll(1, "units")

        def _bind_spin_controls_mousewheel(_event: tk.Event) -> None:
            scroll_canvas.bind_all("<MouseWheel>", _wheel_spin_controls)
            scroll_canvas.bind_all("<Button-4>", _wheel_spin_controls_up)
            scroll_canvas.bind_all("<Button-5>", _wheel_spin_controls_down)

        def _unbind_spin_controls_mousewheel(_event: tk.Event) -> None:
            try:
                scroll_canvas.unbind_all("<MouseWheel>")
                scroll_canvas.unbind_all("<Button-4>")
                scroll_canvas.unbind_all("<Button-5>")
            except tk.TclError:
                pass

        scroll_outer.bind("<Enter>", _bind_spin_controls_mousewheel)
        scroll_outer.bind("<Leave>", _unbind_spin_controls_mousewheel)

        inner = tk.Frame(body, bg=PRIZE_BOARD_CONTENT_BG)
        inner.pack(fill=tk.BOTH, expand=True, padx=18, pady=(0, 18))
        app._build_spin_controls_inner(inner)

        def _initial_scroll_sync() -> None:
            _sync_scroll_body_width()
            try:
                scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))
            except tk.TclError:
                pass

        app.after_idle(_initial_scroll_sync)
        app.after_idle(app._on_spin_controls_wheel_panel_ready)
        app.after_idle(app._refresh_fill_skipped_hint)

        app._spin_controls_wheel_link_reset()
        self.bind("<Configure>", self._on_spin_controls_configure)

    def _on_spin_controls_configure(self, event: tk.Event) -> None:
        if event.widget is not self:
            return
        self._app._schedule_spin_controls_wheel_sync()

    def _on_close(self) -> None:
        self._app._cancel_spin_controls_wheel_sync()
        self._app._wheel_control_link_ref = None
        self._app._controls_win = None
        try:
            c = getattr(self, "_spin_controls_scroll_canvas", None)
            if c is not None:
                c.unbind_all("<MouseWheel>")
                c.unbind_all("<Button-4>")
                c.unbind_all("<Button-5>")
        except tk.TclError:
            pass
        try:
            self.destroy()
        except tk.TclError:
            pass


class DrawPrizeApp(tk.Tk):
    WHEEL_CELL = int(round(118 * 1.5 * 0.8 * 0.9 * WHEEL_VERTICAL_SCALE))
    WHEEL_STRIP_LEN = 44
    # Spin strip: multiple packed laps (more prizes on screen + longer travel).
    WHEEL_SPIN_LAP_LEN = 32
    WHEEL_SPIN_LAPS = 4
    WHEEL_SPIN_CELL_GAP = 5.0
    # Home / idle horizontal strip: logical preview length; physical strip is doubled for seamless carousel.
    WHEEL_IDLE_STRIP_LEN = 11
    WHEEL_IDLE_CAROUSEL_COPIES = 2
    # Short strip while the worker loads picks (before the full spin strip is built).
    WHEEL_LOADING_STRIP_LEN = 15
    # Total horizontal clear pixels between adjacent wheel item rectangles (scaled with cell).
    WHEEL_CELL_INTER_GAP = int(round(18 * 1.5 * 0.8 * 0.9 * 1.1))
    # Strip spin: duration + ease-out so the wheel rolls fast then progressively slows into the winner.
    WHEEL_SPIN_DURATION_FAST_SEC = 2.0
    WHEEL_SPIN_DURATION_SLOW_SEC = 4.5
    WHEEL_SPIN_EASE_OUT_POWER = 5.0
    WHEEL_SPIN_TICK_MS = 16
    # Idle home strip: slow scroll left (loops when it reaches end of usable range).
    WHEEL_IDLE_DRIFT_MS = 45
    WHEEL_IDLE_DRIFT_PX_PER_SEC = 33.0
    # Resizing "Spin & controls" window stretches wheel strip (linked delta from sizes at open).
    WHEEL_CANVAS_H_LINK_MIN = int(round(96 * WHEEL_VERTICAL_SCALE))
    WHEEL_CANVAS_H_LINK_MAX = int(round(540 * WHEEL_VERTICAL_SCALE))
    # Main overlay: winner session listbox shows at most this many rows; scroll for the rest.
    OVERLAY_WINNER_LIST_VISIBLE_ROWS = 25
    TEXAS_TZ = ZoneInfo("America/Chicago")

    @classmethod
    def _wheel_cell_xmargins(cls, *, dense: bool = False) -> tuple[int, int]:
        g = max(10, cls.WHEEL_CELL_INTER_GAP // 2) if dense else cls.WHEEL_CELL_INTER_GAP
        return g // 2, g - g // 2

    @classmethod
    def _wheel_slot_center_offset(cls, *, dense: bool = False) -> float:
        """X offset from i*cw - scroll to the visual center of slot i (matches Tk (x0+x1)/2)."""
        cw = float(cls.WHEEL_CELL)
        pl, pr = cls._wheel_cell_xmargins(dense=dense)
        return (float(pl) + cw - float(pr)) / 2.0

    @classmethod
    def _wheel_scroll_to_center_index(
        cls, index: float, viewport_w: float, *, dense: bool = False
    ) -> float:
        """Scroll value so slot ``index`` is centered under x = viewport_w/2."""
        cw = float(cls.WHEEL_CELL)
        return index * cw + cls._wheel_slot_center_offset(dense=dense) - viewport_w / 2.0

    @classmethod
    def _wheel_spin_ease(cls, t: float) -> float:
        """
        Ease-out progress for spin animation (0 = start, 1 = landed).

        High power = most distance early, then a long progressive slowdown at the end.
        """
        t = max(0.0, min(1.0, float(t)))
        if t >= 1.0:
            return 1.0
        omt = 1.0 - t
        return 1.0 - omt ** float(cls.WHEEL_SPIN_EASE_OUT_POWER)

    def _wheel_primary_viewport_w(self) -> float:
        """Width used when storing scroll (OBS / hidden strip canvas), not Spin & controls preview."""
        try:
            return max(2.0, float(int(self.wheel_canvas.winfo_width())))
        except tk.TclError:
            return 1000.0

    def _wheel_reference_viewport_w(self) -> float:
        w = getattr(self, "_wheel_ref_viewport_w", None)
        if w is None or float(w) < 2.0:
            return self._wheel_primary_viewport_w()
        return float(w)

    def _wheel_note_scroll_viewport(self, viewport_w: float | None = None) -> None:
        if viewport_w is None:
            viewport_w = self._wheel_primary_viewport_w()
        self._wheel_ref_viewport_w = max(2.0, float(viewport_w))

    def _wheel_scroll_convert_viewport(self, scroll: float, from_w: float, to_w: float) -> float:
        """Same slot under the pointer after resize: convert scroll between viewport widths."""
        if abs(from_w - to_w) < 0.5:
            return float(scroll)
        mid = float(self._wheel_slot_center_offset())
        cw = float(self.WHEEL_CELL)
        index = (float(scroll) + from_w / 2.0 - mid) / cw
        return index * cw + mid - to_w / 2.0

    def _wheel_scroll_for_canvas(self, c: tk.Canvas, scroll_at_ref: float) -> float:
        ref = self._wheel_reference_viewport_w()
        try:
            w = max(2.0, float(int(c.winfo_width())))
        except tk.TclError:
            return float(scroll_at_ref)
        return self._wheel_scroll_convert_viewport(float(scroll_at_ref), ref, w)

    def _wheel_spin_duration_sec(self) -> float:
        mode = "fast"
        var = getattr(self, "wheel_spin_speed_var", None)
        if var is not None:
            try:
                mode = str(var.get() or "fast").strip().lower()
            except tk.TclError:
                pass
        if mode == "slow":
            return float(type(self).WHEEL_SPIN_DURATION_SLOW_SEC)
        return float(type(self).WHEEL_SPIN_DURATION_FAST_SEC)

    def _wheel_scroll_for_strip_render_on_canvas(self, c: tk.Canvas) -> float:
        s = self._wheel_scroll_for_canvas(c, float(self._wheel_scroll))
        if self._wheel_idle_drift_active():
            ref = self._wheel_reference_viewport_w()
            try:
                w = max(2.0, float(int(c.winfo_width())))
            except tk.TclError:
                w = ref
            off = float(getattr(self, "_wheel_idle_offset", 0.0) or 0.0)
            if ref > 0:
                s += off * (w / ref)
            else:
                s += off
        return s

    def __init__(self) -> None:
        super().__init__()
        self._void_bg_widgets: list[tk.Misc] = []
        self.title("Energy Break — Overlay")
        self.geometry(f"{OVERLAY_WIDTH}x{OVERLAY_HEIGHT_COMPACT}")
        self.minsize(380, 200)
        self._overlay_compact_height = OVERLAY_HEIGHT_COMPACT
        self._prize_win: PrizeListOverlay | None = None
        self._controls_win: SpinControlsWindow | None = None
        self._spin_controls_wheel_sync_after: str | int | None = None
        self._wheel_control_link_ref: tuple[int, int] | None = None
        self._prize_board_totals_label: tk.Label | None = None
        self.configure(bg=self._void_bg())
        try:
            self.attributes("-topmost", False)
        except tk.TclError:
            pass
        self.overrideredirect(True)

        self._drag_offset: tuple[int, int] = (0, 0)
        self._session: draw_prize.FileDrawSession | None = None
        self._session_target = ""
        self._anim_after: str | int | None = None
        self._pulse_after: str | int | None = None
        self._pending_super: tuple[draw_prize.FileDrawSession, draw_prize.SpinResult] | None = None
        self._super_reroll_used = False
        self._busy = False
        self._obs_last_result_spot: int | None = None
        self._obs_last_result_sku: str = ""
        self._backfill_target_spot: int | None = None
        self._winner_log_path: Path | None = None
        self._winner_next_spin = 1
        self._active_wheel_preset_id = ""
        self._undo_spin_snap: _UndoSpinSnapshot | None = None
        self._winner_overlay_poll_after: str | int | None = None
        self._winner_overlay_poll_started = False
        self._winner_overlay_last_sig: tuple[str, float] | None = None

        self._title_font = tkfont.Font(self, family="Segoe UI", size=13, weight="bold")
        self._wheel_font = tkfont.Font(self, family="Segoe UI", size=10, weight="normal")
        self._wheel_font_bold = tkfont.Font(self, family="Segoe UI", size=10, weight="bold")
        self._btn_font = tkfont.Font(self, family="Segoe UI", size=12, weight="bold")

        self._wheel_strip: list[str] = []
        self._wheel_win_idx = 0
        self._wheel_target_scroll = 0.0
        self._wheel_scroll = 0.0
        self._wheel_ref_viewport_w = 1000.0
        self._wheel_idle_offset = 0.0
        self._wheel_idle_drift_after: str | int | None = None
        self._wheel_pulse_highlight = False
        self._wheel_dense_pack_active = False
        self._customer_empty_wheel = False
        self._wheel_sku_to_img: dict[str, str] = {}
        self._wheel_image_cache: dict[str, object] = {}
        self._wheel_http_server: object | None = None
        self._wheel_http_port: int = 0
        self._wheel_html_snapshot_cache: dict = {}
        self._wheel_banner_title: str | None = None
        self._wheel_banner_subtitle: str | None = None
        self._wheel_banner_is_error: bool = False
        self._inventory_image_cache: dict[tuple[str, int], object] = {}
        self._prize_board_grid_cols = tk.IntVar(value=PRIZE_BOARD_GRID_COLS_DEFAULT)
        self._prize_board_tile_rows_fit = tk.IntVar(value=PRIZE_BOARD_ROWS_FIT_DEFAULT)
        self.wheel_preset_var = tk.StringVar(value="")
        self._prize_board_resize_after: str | int | None = None
        self._prize_board_last_layout: tuple[int, int, int] | None = None
        self._app_state_save_after: str | int | None = None

        self._prize_board_grid_cols.trace_add("write", self._on_prize_board_grid_var_write)
        self._prize_board_tile_rows_fit.trace_add("write", self._on_prize_board_grid_var_write)

        self._build_drag_bar()

        self._setup_expanded = tk.BooleanVar(value=False)
        self._log_visible = tk.BooleanVar(value=False)

        self._setup_toggle_frame = tk.Frame(self, bg=self._void_bg())
        self._setup_toggle_frame.pack(fill=tk.X, padx=8, pady=(0, 4))
        self._register_void_bg(self._setup_toggle_frame)
        self._chk_setup_expand = tk.Checkbutton(
            self._setup_toggle_frame,
            text="Show setup (choose files & options)",
            variable=self._setup_expanded,
            command=self._toggle_setup,
            bg=self._void_bg(),
            fg=WHEEL_FG,
            selectcolor=WHEEL_CELL_BG,
            activebackground=self._void_bg(),
            activeforeground=WHEEL_FG,
            font=("Segoe UI", 10),
        )
        self._chk_setup_expand.pack(side=tk.LEFT)
        self._register_void_bg(self._chk_setup_expand)

        self.setup_frame = tk.Frame(self, bg=self._void_bg())
        self._register_void_bg(self.setup_frame)
        self._build_setup_scroll_shell()
        self._build_setup_form(self._setup_scroll_body)
        if not self._load_app_state():
            self._init_winner_session_log()

        self._wheel_host = tk.Frame(self, bg=self._void_bg())
        self._register_void_bg(self._wheel_host)
        self._wheel_host.pack(fill=tk.X, expand=False, padx=8, pady=(4, 8))
        self._build_wheel_area(self._wheel_host)
        self._build_prize_list_window()

        self._update_spin_counter_label()
        self._log(
            f"Chroma (wheel only): empty chrome uses solid {OBS_CHROMA_KEY} — key that out in OBS if you capture the wheel. "
            "Do not use the same green in prize images.\n"
        )
        self._log(
            "Spin & controls opens automatically in a separate window. "
            "Prize board: use the button on the wheel bar. "
            "HTML wheel for OBS: Spin & controls → Open HTML wheel. "
            "Drag the orange bar to move, ✕ to exit.\n"
        )
        self._reset_wheel_idle()
        self._refresh_prizes_label()
        self._sync_obs_overlay()
        self._wheel_html_server_start()
        self._wheel_publish_html_snapshot()
        self.after_idle(self._show_spin_controls)

    def _on_prize_board_grid_var_write(self, *_args: object) -> None:
        self._schedule_prize_board_layout_refresh()

    def _void_bg(self) -> str:
        """Background for overlay chrome keyed out in OBS (solid green, not OS transparency)."""
        return OBS_CHROMA_KEY

    def _prize_board_surface_bg(self) -> str:
        """Panel background for the on-monitor prize board (not chroma green)."""
        return PRIZE_BOARD_CONTENT_BG

    def _prize_board_cols_rows_clamped(self) -> tuple[int, int]:
        """User grid: columns (tiles per row) and rows (used only to fit tile height to the panel)."""
        try:
            c = int(self._prize_board_grid_cols.get())
        except (tk.TclError, ValueError, TypeError):
            c = PRIZE_BOARD_GRID_COLS_DEFAULT
        c = max(2, min(24, c))
        try:
            r = int(self._prize_board_tile_rows_fit.get())
        except (tk.TclError, ValueError, TypeError):
            r = PRIZE_BOARD_ROWS_FIT_DEFAULT
        r = max(2, min(20, r))
        return c, r

    def _prize_board_slot_thumb_for_inner(self, ri: tk.Frame) -> tuple[int, int, int]:
        """Square slot size and thumbnail side from inner frame size; returns (slot_px, thumb_px, cols)."""
        cols, rows_fit = self._prize_board_cols_rows_clamped()
        try:
            ri.update_idletasks()
        except tk.TclError:
            pass
        g = INV_GRID_GAP
        edge = 16
        w = max(260, int(ri.winfo_width()))
        h = max(140, int(ri.winfo_height()))
        slot_w = (w - edge * 2 - cols * (2 * g)) // max(1, cols)
        slot_h = (h - edge * 2 - rows_fit * (2 * g)) // max(1, rows_fit)
        slot = min(slot_w, slot_h)
        slot = max(PRIZE_BOARD_SLOT_MIN_PX, min(PRIZE_BOARD_SLOT_MAX_PX, slot))
        thumb = max(22, min(int(round(slot * 0.62)), slot - 16))
        return slot, thumb, cols

    def _schedule_prize_board_layout_refresh(self, _ev: object | None = None) -> None:
        """Debounce grid rebuild when the prize board is resized or grid prefs change."""
        aid = getattr(self, "_prize_board_resize_after", None)
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
        self._prize_board_resize_after = self.after(100, self._prize_board_layout_refresh_cb)

    def _prize_board_layout_refresh_cb(self) -> None:
        self._prize_board_resize_after = None
        self._refresh_remaining_skus_panel()

    def _cancel_prize_board_resize_refresh(self) -> None:
        aid = getattr(self, "_prize_board_resize_after", None)
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
            self._prize_board_resize_after = None

    def _register_void_bg(self, w: tk.Misc) -> None:
        if w not in self._void_bg_widgets:
            self._void_bg_widgets.append(w)

    def _sync_obs_overlay(self) -> None:
        void = self._void_bg()
        try:
            self.configure(bg=void)
        except tk.TclError:
            pass
        for w in self._void_bg_widgets:
            try:
                if isinstance(w, tk.Checkbutton):
                    w.configure(bg=void, activebackground=void)
                else:
                    w.configure(bg=void)
            except tk.TclError:
                pass
        if os.name == "nt":
            try:
                self.attributes("-transparentcolor", "")
            except tk.TclError:
                pass
            pw = getattr(self, "_prize_win", None)
            if pw is not None:
                try:
                    if pw.winfo_exists():
                        pw.attributes("-transparentcolor", "")
                except tk.TclError:
                    pass

    def _open_html_wheel_in_browser(self) -> None:
        """Open the HTML wheel page (same feed as OBS Browser Source)."""
        port = int(getattr(self, "_wheel_http_port", 0) or 0)
        if port <= 0:
            messagebox.showinfo(
                "HTML wheel",
                "The local wheel server did not start. Ensure web/wheel_spin.html exists next to the app.",
            )
            return
        url = f"http://127.0.0.1:{port}/"
        try:
            webbrowser.open(url)
        except OSError as e:
            messagebox.showerror("HTML wheel", f"Could not open browser:\n{e}\n\nURL:\n{url}")

    def _wheel_shutdown_html_server(self) -> None:
        httpd = getattr(self, "_wheel_http_server", None)
        if httpd is None:
            return
        self._wheel_http_server = None
        self._wheel_http_port = 0
        try:
            httpd.shutdown()
        except Exception:
            pass

    def _wheel_html_server_start(self) -> None:
        self._wheel_shutdown_html_server()
        html_path = resource_root() / "web" / "wheel_spin.html"
        try:
            from wheel_html_server import start_wheel_html_server as _wheel_http_boot
        except ImportError:
            self._log("HTML wheel: could not import wheel_html_server.py (browser mirror disabled).\n")
            return

        def _snap() -> dict:
            return dict(getattr(self, "_wheel_html_snapshot_cache", {}))

        def _img(sku: str) -> tuple[bytes, str] | None:
            return self._wheel_html_image_bytes(sku)

        try:
            from wheel_html_server import WHEEL_HTML_SERVER_PORT as _html_port_default
        except ImportError:
            _html_port_default = 8765
        httpd, port, warn = _wheel_http_boot(html_path, _snap, _img)
        if httpd is None or not port:
            self._log("HTML wheel: web/wheel_spin.html not found — browser / OBS mirror disabled.\n")
            return
        self._wheel_http_server = httpd
        self._wheel_http_port = int(port)
        url = f"http://127.0.0.1:{port}/"
        if port == _html_port_default:
            obs_hint = f"OBS Browser Source (fixed URL): {url}"
        else:
            obs_hint = f"HTML wheel (OBS): {url}"
        self._log(
            f"{obs_hint}  — transparent page; strip width "
            f"{int(round(100 * HTML_WHEEL_DISPLAY_SCALE * HTML_WHEEL_STRIP_WIDTH_MUL))}% of app cells, "
            f"strip vertical pad +{int(round(100 * (HTML_WHEEL_STRIP_PAD_V_MUL - 1)))}% (after display scale).\n"
        )
        if warn:
            self._log(f"HTML wheel warning: {warn}\n")

    def _wheel_html_image_bytes(self, sku: str) -> tuple[bytes, str] | None:
        if sku not in self._wheel_strip and sku not in self._wheel_sku_to_img:
            return None
        ref = self._wheel_sku_to_img.get(sku, "")
        refn = _normalize_image_ref(ref)
        if not refn or refn.lower().startswith(("http://", "https://")):
            return None
        path = self._resolve_prize_image_path(refn)
        if path is None or not path.is_file():
            return None
        try:
            blob = path.read_bytes()
        except OSError:
            return None
        ct, _enc = mimetypes.guess_type(str(path))
        return blob, ct or "application/octet-stream"

    def _clear_obs_wheel_result(self) -> None:
        self._obs_last_result_spot = None
        self._obs_last_result_sku = ""

    def _set_obs_wheel_result(self, spot: int, sku: str) -> None:
        self._obs_last_result_spot = int(spot)
        self._obs_last_result_sku = (sku or "").strip()

    def _wheel_obs_overlay_lines(
        self,
        spot_num: int | None,
        landed: bool,
        winning_sku: str,
        strip_len: int,
    ) -> tuple[str, str]:
        """Primary + secondary lines for the HTML / OBS spot chip (not the Tk wheel status)."""
        spot_x = f"Spot {spot_num}" if spot_num is not None else "Spot —"
        WSL = int(type(self).WHEEL_STRIP_LEN)
        WIL = int(type(self).WHEEL_IDLE_STRIP_LEN)
        anim = self._anim_after is not None
        pending_super = self._pending_super is not None
        busy = self._busy
        last_spot = self._obs_last_result_spot
        last_sku = (self._obs_last_result_sku or "").strip()
        win_sku = (winning_sku or "").strip()

        if strip_len == 0:
            return spot_x, ""

        bf = getattr(self, "_backfill_target_spot", None)
        if anim or (
            busy
            and not pending_super
            and strip_len not in (WSL, WIL, WIL * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES))
        ):
            if bf is not None:
                return f"Spinning to fill Spot {bf} (was skipped)", ""
            return f"Spinning for {spot_x}", ""

        if pending_super:
            sub = win_sku
            if not sub and self._pending_super is not None:
                sub = (self._pending_super[1].sku or "").strip()
            line2 = f"Landed: {sub}" if sub else "Choose REROLL or KEEP"
            if self._super_reroll_used:
                return f"Super — {spot_x} · final pick", line2
            return f"Super — {spot_x}", line2

        if strip_len == WSL and landed:
            if last_spot is not None:
                return f"Winner Spot {last_spot}", last_sku or win_sku
            if busy and spot_num is not None:
                if bf is not None:
                    return f"Winner Spot {bf}", win_sku or last_sku
                return f"Winner {spot_x}", win_sku or last_sku
            if spot_num is not None and win_sku:
                return f"Winner {spot_x}", win_sku
            if spot_num is not None:
                return f"Winner {spot_x}", ""
            return spot_x, win_sku

        if strip_len == WIL or strip_len == WIL * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES):
            return f"Next draw: {spot_x}", ""

        if strip_len == int(type(self).WHEEL_LOADING_STRIP_LEN):
            return f"Spinning for {spot_x}", ""

        return spot_x, ""

    def _wheel_prizes_left_label_text(self) -> str:
        """Short line for wheel chrome + HTML mirror (same quantity source as the main Total prizes label)."""
        n = self._total_qty_remaining()
        if n is None:
            return "Total prizes left: —"
        return f"Total prizes left: {n}"

    def _wheel_build_html_snapshot(self) -> dict:
        colors = {
            "slotBg": WHEEL_STRIP_SLOT_BG,
            "slotBd": WHEEL_STRIP_SLOT_BD,
            "fg": WHEEL_STRIP_FG,
            "winBg": WHEEL_STRIP_WIN_BG,
            "winFg": WHEEL_STRIP_WIN_FG,
            "winRing": WHEEL_STRIP_WIN_RING,
            "pointer": WHEEL_STRIP_POINTER,
            "pointerEdge": WHEEL_STRIP_POINTER_EDGE,
            "track": WHEEL_STRIP_TRACK,
            "title": WHEEL_TITLE,
            "accent": WHEEL_ACCENT,
        }
        spot = "Spot 1"
        st = ""
        if hasattr(self, "spin_counter_label"):
            try:
                spot = str(self.spin_counter_label.cget("text") or spot)
            except tk.TclError:
                pass
        if hasattr(self, "wheel_status"):
            try:
                st = str(self.wheel_status.cget("text") or "")
            except tk.TclError:
                pass
        spot_num = self._parse_spot_number_from_label(spot)
        if getattr(self, "_anim_after", None) is not None:
            spot_assignment = ""
        else:
            spot_assignment = (
                self._winner_assignment_display_for_spot(spot_num)
                if spot_num is not None
                else ""
            )
        base: dict = {
            "chroma": OBS_CHROMA_KEY,
            "colors": colors,
            "spotLabel": spot,
            "spotNumber": spot_num,
            "spotAssignment": spot_assignment,
            "statusText": st,
            "prizesLeftLabel": self._wheel_prizes_left_label_text(),
        }
        if self._wheel_banner_title and not self._wheel_strip:
            oh, os_ = self._wheel_obs_overlay_lines(spot_num, False, "", 0)
            base["obsHeadline"] = oh or spot
            base["obsSubhead"] = os_ or (spot_assignment or "")
            base.update(
                {
                    "mode": "banner",
                    "bannerTitle": self._wheel_banner_title,
                    "bannerSubtitle": self._wheel_banner_subtitle or "",
                    "bannerAccent": "error" if self._wheel_banner_is_error else "muted",
                }
            )
            return base

        c = self.wheel_canvas
        try:
            h = int(c.cget("height"))
            w = max(int(c.winfo_width()), 2)
        except tk.TclError:
            h, w = 120, 520
        cx = w / 2.0
        scroll_model = float(self._wheel_scroll)
        scroll_vis = self._wheel_scroll_for_strip_render()
        cw = float(self.WHEEL_CELL)
        dense = bool(getattr(self, "_wheel_dense_pack_active", False))
        pl, pr = self._wheel_cell_xmargins(dense=dense)
        gap = float(type(self).WHEEL_SPIN_CELL_GAP if dense else WHEEL_STRIP_CELL_GAP)
        hs = float(HTML_WHEEL_DISPLAY_SCALE)
        hw = hs * float(HTML_WHEEL_STRIP_WIDTH_MUL)
        pv = hs * float(HTML_WHEEL_STRIP_PAD_V_MUL)
        hch = float(HTML_WHEEL_STRIP_CELL_HEIGHT_MUL)
        landed = abs(scroll_model - float(self._wheel_target_scroll)) < 1.5
        slot_mid = self._wheel_slot_center_offset(dense=dense)
        cells: list[dict] = []
        winning_sku = ""
        idle_carousel = len(self._wheel_strip) == int(type(self).WHEEL_IDLE_STRIP_LEN) * int(
            type(self).WHEEL_IDLE_CAROUSEL_COPIES
        )
        for i, raw in enumerate(self._wheel_strip):
            cell_center_x = i * cw + slot_mid - scroll_vis
            if idle_carousel:
                win_cell = False
            elif len(self._wheel_strip) == self.WHEEL_IDLE_STRIP_LEN:
                win_cell = landed and abs(cell_center_x - cx) < cw * 0.38
            else:
                win_cell = landed and i == self._wheel_win_idx and abs(cell_center_x - cx) < cw * 0.38
            if win_cell:
                winning_sku = self._wheel_label_text(raw)
            ref = self._wheel_sku_to_img.get(raw, "")
            refn = _normalize_image_ref(ref)
            if refn and refn.lower().startswith(("http://", "https://")):
                img_url = refn
            elif refn:
                img_url = f"/wheel/img?sku={quote(raw, safe='')}"
            else:
                img_url = ""
            cells.append(
                {
                    "sku": self._wheel_label_text(raw),
                    "img": img_url,
                    "win": bool(win_cell),
                }
            )
        oh, os_ = self._wheel_obs_overlay_lines(spot_num, landed, winning_sku, len(self._wheel_strip))
        base["obsHeadline"] = oh or spot
        base["obsSubhead"] = os_
        base.update(
            {
                "mode": "strip",
                "viewportW": w,
                "canvasH": h,
                "cellW": int(round(self.WHEEL_CELL * hw)),
                "scroll": scroll_vis * hw,
                "padTop": max(18, int(round(WHEEL_STRIP_PAD_TOP * pv / hch))),
                "padBottom": max(8, int(round(WHEEL_STRIP_PAD_BOTTOM * pv / hch))),
                "gap": gap * hw,
                "pl": int(round(pl * hw)),
                "pr": int(round(pr * hw)),
                "pulse": bool(self._wheel_pulse_highlight),
                "cells": cells,
            }
        )
        return base

    def _wheel_html_snapshot_spin_frame(self, prev: dict) -> dict | None:
        """During strip spin animation, reuse cell image/SKU payload and only update scroll + win flags."""
        if not self._wheel_strip:
            return None
        prev_cells = prev.get("cells")
        if not isinstance(prev_cells, list) or len(prev_cells) != len(self._wheel_strip):
            return None
        for i, raw in enumerate(self._wheel_strip):
            pc = prev_cells[i]
            if not isinstance(pc, dict):
                return None
            if (pc.get("sku") or "") != self._wheel_label_text(raw):
                return None
        c = self.wheel_canvas
        try:
            h = int(c.cget("height"))
            w = max(int(c.winfo_width()), 2)
        except tk.TclError:
            return None
        if w != int(prev.get("viewportW") or -1) or h != int(prev.get("canvasH") or -1):
            return None
        cx = w / 2.0
        scroll = float(self._wheel_scroll)
        cw = float(self.WHEEL_CELL)
        dense_sf = bool(getattr(self, "_wheel_dense_pack_active", False))
        pl, pr = self._wheel_cell_xmargins(dense=dense_sf)
        slot_mid = self._wheel_slot_center_offset(dense=dense_sf)
        landed = abs(scroll - float(self._wheel_target_scroll)) < 1.5
        hs = float(HTML_WHEEL_DISPLAY_SCALE)
        hw = hs * float(HTML_WHEEL_STRIP_WIDTH_MUL)
        winning_sku = ""
        new_cells: list[dict] = []
        for i, raw in enumerate(self._wheel_strip):
            pc = prev_cells[i]
            cell_center_x = i * cw + slot_mid - scroll
            win_cell = landed and i == self._wheel_win_idx and abs(cell_center_x - cx) < cw * 0.38
            if win_cell:
                winning_sku = self._wheel_label_text(raw)
            nd = dict(pc)
            nd["win"] = bool(win_cell)
            new_cells.append(nd)
        spot = "Spot 1"
        if hasattr(self, "spin_counter_label"):
            try:
                spot = str(self.spin_counter_label.cget("text") or spot)
            except tk.TclError:
                pass
        spot_num = self._parse_spot_number_from_label(spot)
        oh, os_ = self._wheel_obs_overlay_lines(
            spot_num, landed, winning_sku, len(self._wheel_strip)
        )
        out = dict(prev)
        out["scroll"] = scroll * hw
        out["cells"] = new_cells
        out["pulse"] = bool(self._wheel_pulse_highlight)
        out["obsHeadline"] = oh or spot
        out["obsSubhead"] = os_
        out["spotLabel"] = spot
        out["spotNumber"] = spot_num
        out["spotAssignment"] = ""
        out["prizesLeftLabel"] = self._wheel_prizes_left_label_text()
        return out

    def _wheel_publish_html_snapshot(self) -> None:
        try:
            prev = getattr(self, "_wheel_html_snapshot_cache", None)
            if (
                getattr(self, "_anim_after", None) is not None
                and isinstance(prev, dict)
                and prev.get("mode") == "strip"
            ):
                fast = self._wheel_html_snapshot_spin_frame(prev)
                if fast is not None:
                    self._wheel_html_snapshot_cache = fast
                    return
            self._wheel_html_snapshot_cache = self._wheel_build_html_snapshot()
        except Exception as e:
            try:
                self._log(f"Wheel snapshot error: {type(e).__name__}: {e}\n")
            except Exception:
                pass
            try:
                pl_err = self._wheel_prizes_left_label_text()
            except Exception:
                pl_err = "Total prizes left: —"
            self._wheel_html_snapshot_cache = {
                "mode": "banner",
                "chroma": OBS_CHROMA_KEY,
                "colors": {},
                "spotLabel": "",
                "spotNumber": None,
                "spotAssignment": "",
                "obsHeadline": "",
                "obsSubhead": "",
                "statusText": "",
                "bannerTitle": "Wheel snapshot error",
                "bannerSubtitle": "",
                "bannerAccent": "error",
                "prizesLeftLabel": pl_err,
            }

    def _build_spin_controls_window(self) -> None:
        if self._controls_win is not None:
            try:
                if self._controls_win.winfo_exists():
                    return
            except tk.TclError:
                pass
            self._controls_win = None
        self._controls_win = SpinControlsWindow(self)

    def _show_spin_controls(self) -> None:
        cw = self._controls_win
        if cw is not None:
            try:
                if cw.winfo_exists():
                    cw.lift()
                    self._on_spin_controls_wheel_panel_ready()
                    return
            except tk.TclError:
                self._controls_win = None
        self._controls_win = SpinControlsWindow(self)

    def _build_prize_list_window(self) -> None:
        if self._prize_win is not None:
            try:
                if self._prize_win.winfo_exists():
                    return
            except tk.TclError:
                pass
            self._prize_win = None
        self._prize_win = PrizeListOverlay(self)

    def _show_prize_board(self) -> None:
        pw = self._prize_win
        if pw is not None:
            try:
                if pw.winfo_exists():
                    pw.lift()
                    return
            except tk.TclError:
                self._prize_win = None
        self._prize_win = PrizeListOverlay(self)
        self._sync_obs_overlay()

    def destroy(self) -> None:
        self._cancel_prize_board_resize_refresh()
        pw = self._prize_win
        if pw is not None:
            try:
                pw.destroy()
            except tk.TclError:
                pass
            self._prize_win = None
        cw = self._controls_win
        if cw is not None:
            try:
                cw.destroy()
            except tk.TclError:
                pass
            self._controls_win = None
        self._cancel_spin_controls_wheel_sync()
        self._wheel_control_link_ref = None
        self.remaining_scroll = None  # type: ignore[assignment]
        self.remaining_canvas = None  # type: ignore[assignment]
        self.remaining_inner = None  # type: ignore[assignment]
        self._remaining_canvas_win = None  # type: ignore[assignment]
        self._prize_board_totals_label = None
        self._save_app_state()
        self._wheel_shutdown_html_server()
        super().destroy()

    def _build_drag_bar(self) -> None:
        self._drag_bar = tk.Frame(self, bg=DRAG_BAR, height=28)
        self._drag_bar.pack(fill=tk.X)
        tk.Button(
            self._drag_bar,
            text=" ✕ ",
            command=self.destroy,
            bg="#c0392b",
            fg="white",
            activebackground="#e74c3c",
            font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT,
            padx=10,
        ).pack(side=tk.RIGHT, padx=4, pady=2)
        drag_lbl = tk.Label(
            self._drag_bar,
            text="  ◎ Energy Break — drag to move  ",
            bg=DRAG_BAR,
            fg="#1a1a1a",
            font=("Segoe UI", 10, "bold"),
        )
        drag_lbl.pack(side=tk.LEFT)
        for w in (self._drag_bar, drag_lbl):
            w.bind("<ButtonPress-1>", self._drag_start)
            w.bind("<B1-Motion>", self._drag_motion)

    def _drag_start(self, event: tk.Event) -> None:
        self._drag_offset = (event.x_root - self.winfo_rootx(), event.y_root - self.winfo_rooty())

    def _drag_motion(self, event: tk.Event) -> None:
        x = event.x_root - self._drag_offset[0]
        y = event.y_root - self._drag_offset[1]
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}+{x}+{y}")

    def _toggle_setup(self) -> None:
        if self._setup_expanded.get():
            self.setup_frame.pack(fill=tk.X, padx=8, pady=(0, 4), before=self._wheel_host)
            self.after_idle(self._overlay_resize_for_setup)
        else:
            self.setup_frame.pack_forget()
            self.after_idle(self._overlay_resize_compact)

    def _toggle_log(self) -> None:
        if self._log_visible.get():
            self.log_frame.pack(fill=tk.BOTH, expand=False, pady=(4, 0))
        else:
            self.log_frame.pack_forget()
        self._setup_sync_scroll_region()
        if self._setup_expanded.get():
            self.after_idle(self._overlay_resize_for_setup)

    def _build_setup_scroll_shell(self) -> None:
        """Scrollable setup panel so Step 2, defaults, and log are not clipped on the overlay."""
        vb = self._void_bg()
        host = tk.Frame(self.setup_frame, bg=vb)
        host.pack(fill=tk.BOTH, expand=True)
        self._register_void_bg(host)

        tk.Label(
            host,
            text="Scroll ↓ for Step 2, wheel preset, and options",
            font=("Segoe UI", 8),
            bg=vb,
            fg=WHEEL_MUTED,
            anchor=tk.W,
        ).pack(fill=tk.X, padx=2, pady=(0, 4))

        scroll_outer = tk.Frame(host, bg=vb)
        scroll_outer.pack(fill=tk.BOTH, expand=True)
        self._setup_scroll_outer = scroll_outer

        self._setup_scroll_canvas = tk.Canvas(
            scroll_outer,
            bg=vb,
            height=OVERLAY_SETUP_SCROLL_MAX,
            highlightthickness=0,
            bd=0,
        )
        self._setup_scrollbar = tk.Scrollbar(
            scroll_outer,
            orient=tk.VERTICAL,
            command=self._setup_scroll_canvas.yview,
            bg=vb,
            troughcolor=WHEEL_CELL_BG,
            activebackground=WHEEL_POINTER,
        )
        self._setup_scroll_canvas.configure(yscrollcommand=self._setup_scrollbar.set)
        self._setup_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._setup_scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._setup_scroll_body = tk.Frame(self._setup_scroll_canvas, bg=vb)
        self._setup_scroll_body_win = self._setup_scroll_canvas.create_window(
            (0, 0), window=self._setup_scroll_body, anchor=tk.NW
        )

        def _sync_setup_scroll_width(_event: object | None = None) -> None:
            try:
                cw = max(200, int(self._setup_scroll_canvas.winfo_width()) - 4)
            except tk.TclError:
                return
            try:
                self._setup_scroll_canvas.itemconfigure(self._setup_scroll_body_win, width=cw)
            except tk.TclError:
                pass
            wrap = max(260, cw - 24)
            for lab in getattr(self, "_setup_wrap_labels", ()):
                try:
                    lab.configure(wraplength=wrap)
                except tk.TclError:
                    pass

        def _on_setup_body_configure(_event: tk.Event) -> None:
            self._setup_sync_scroll_region()
            _sync_setup_scroll_width()

        def _on_setup_canvas_configure(_event: tk.Event) -> None:
            _sync_setup_scroll_width()

        self._setup_scroll_body.bind("<Configure>", _on_setup_body_configure, add="+")
        self._setup_scroll_canvas.bind("<Configure>", _on_setup_canvas_configure, add="+")

        def _wheel_setup(event: tk.Event) -> None:
            if getattr(event, "delta", 0):
                self._setup_scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _wheel_setup_up(_event: tk.Event) -> None:
            self._setup_scroll_canvas.yview_scroll(-1, "units")

        def _wheel_setup_down(_event: tk.Event) -> None:
            self._setup_scroll_canvas.yview_scroll(1, "units")

        def _bind_setup_wheel(_event: tk.Event) -> None:
            self._setup_scroll_canvas.bind_all("<MouseWheel>", _wheel_setup)
            self._setup_scroll_canvas.bind_all("<Button-4>", _wheel_setup_up)
            self._setup_scroll_canvas.bind_all("<Button-5>", _wheel_setup_down)

        def _unbind_setup_wheel(_event: tk.Event) -> None:
            try:
                self._setup_scroll_canvas.unbind_all("<MouseWheel>")
                self._setup_scroll_canvas.unbind_all("<Button-4>")
                self._setup_scroll_canvas.unbind_all("<Button-5>")
            except tk.TclError:
                pass

        scroll_outer.bind("<Enter>", _bind_setup_wheel)
        scroll_outer.bind("<Leave>", _unbind_setup_wheel)

        self._setup_wrap_labels: list[tk.Label] = []
        self.after_idle(_sync_setup_scroll_width)

    def _setup_sync_scroll_region(self) -> None:
        c = getattr(self, "_setup_scroll_canvas", None)
        if c is None:
            return
        try:
            self.update_idletasks()
            c.configure(scrollregion=c.bbox("all"))
        except tk.TclError:
            pass

    def _overlay_chrome_height(self, *, include_setup: bool) -> int:
        self.update_idletasks()
        parts = [self._drag_bar, self._setup_toggle_frame, self._wheel_host]
        if include_setup and self._setup_expanded.get():
            try:
                if self.setup_frame.winfo_ismapped():
                    parts.insert(2, self.setup_frame)
            except tk.TclError:
                pass
        total = 12
        for w in parts:
            try:
                if w.winfo_ismapped():
                    total += int(w.winfo_reqheight())
            except tk.TclError:
                continue
        return max(OVERLAY_HEIGHT_COMPACT, total)

    def _overlay_resize_compact(self) -> None:
        h = self._overlay_chrome_height(include_setup=False)
        self._overlay_compact_height = h
        x, y = self.winfo_x(), self.winfo_y()
        try:
            self.geometry(f"{OVERLAY_WIDTH}x{h}+{x}+{y}")
        except tk.TclError:
            pass

    def _overlay_resize_for_setup(self) -> None:
        if not self._setup_expanded.get():
            return
        self._setup_sync_scroll_region()
        try:
            body_h = int(self._setup_scroll_body.winfo_reqheight())
        except tk.TclError:
            body_h = OVERLAY_SETUP_SCROLL_MAX
        view_h = min(max(140, body_h + 8), OVERLAY_SETUP_SCROLL_MAX)
        try:
            self._setup_scroll_canvas.configure(height=view_h)
        except tk.TclError:
            pass
        h = self._overlay_chrome_height(include_setup=True)
        x, y = self.winfo_x(), self.winfo_y()
        try:
            self.geometry(f"{OVERLAY_WIDTH}x{h}+{x}+{y}")
        except tk.TclError:
            pass
        self._setup_sync_scroll_region()

    def _build_setup_form(self, parent: tk.Frame) -> None:
        parent.configure(bg=self._void_bg())
        self._setup_wrap_labels = []
        card_bg = WHEEL_BG
        card_edge = WHEEL_CELL_BORDER
        card = tk.Frame(parent, bg=card_bg, highlightbackground=card_edge, highlightthickness=1)
        card.pack(fill=tk.X, padx=4, pady=(0, 6))
        pad = tk.Frame(card, bg=card_bg)
        pad.pack(fill=tk.X, padx=12, pady=12)

        tk.Label(
            pad,
            text="Where are your files?",
            font=("Segoe UI", 11, "bold"),
            bg=card_bg,
            fg=WHEEL_TITLE,
            anchor=tk.W,
        ).pack(fill=tk.X)
        intro = tk.Label(
            pad,
            text="Pick the prize spreadsheet and the folder that holds your product photos. "
            "Status updates as you type or browse.",
            font=("Segoe UI", 9),
            bg=card_bg,
            fg=WHEEL_MUTED,
            anchor=tk.W,
            justify=tk.LEFT,
            wraplength=400,
        )
        intro.pack(fill=tk.X, pady=(2, 6))
        self._setup_wrap_labels.append(intro)

        def _step_block(
            step_num: int,
            title: str,
            hint: str,
        ) -> tk.Frame:
            block = tk.Frame(pad, bg=card_bg)
            block.pack(fill=tk.X, pady=(0, 12))
            head = tk.Frame(block, bg=card_bg)
            head.pack(fill=tk.X)
            tk.Label(
                head,
                text=f"Step {step_num}",
                font=("Segoe UI", 8, "bold"),
                bg=WHEEL_CELL_BG,
                fg=WHEEL_POINTER,
                padx=6,
                pady=2,
            ).pack(side=tk.LEFT, padx=(0, 8))
            tk.Label(
                head,
                text=title,
                font=("Segoe UI", 10, "bold"),
                bg=card_bg,
                fg=WHEEL_FG,
                anchor=tk.W,
            ).pack(side=tk.LEFT, fill=tk.X)
            hint_lbl = tk.Label(
                block,
                text=hint,
                font=("Segoe UI", 8),
                bg=card_bg,
                fg=WHEEL_MUTED,
                anchor=tk.W,
                justify=tk.LEFT,
                wraplength=400,
            )
            hint_lbl.pack(fill=tk.X, pady=(4, 6))
            self._setup_wrap_labels.append(hint_lbl)
            return block

        def _path_entry_row(block: tk.Frame, textvariable: tk.StringVar) -> tk.Entry:
            row = tk.Frame(block, bg=card_bg)
            row.pack(fill=tk.X)
            ent = tk.Entry(
                row,
                textvariable=textvariable,
                font=("Segoe UI", 10),
                bg=INV_SLOT_FACE,
                fg=WHEEL_FG,
                insertbackground=WHEEL_FG,
                relief=tk.FLAT,
                highlightthickness=1,
                highlightbackground=card_edge,
                highlightcolor=WHEEL_POINTER,
            )
            ent.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)
            return ent

        def _action_buttons(
            block: tk.Frame,
            *,
            pick_cmd: object,
            pick_label: str,
            open_cmd: object,
        ) -> None:
            brow = tk.Frame(block, bg=card_bg)
            brow.pack(fill=tk.X, pady=(8, 0))
            tk.Button(
                brow,
                text=pick_label,
                command=pick_cmd,
                font=("Segoe UI", 9, "bold"),
                bg=BTN_SPIN,
                fg="white",
                activebackground="#3498db",
                activeforeground="white",
                relief=tk.FLAT,
                padx=10,
                pady=6,
                cursor="hand2",
            ).pack(side=tk.LEFT, padx=(0, 8))
            tk.Button(
                brow,
                text="  Open folder  ",
                command=open_cmd,
                font=("Segoe UI", 9),
                bg=WHEEL_CELL_BG,
                fg=WHEEL_FG,
                activebackground=WHEEL_POINTER,
                activeforeground="#1a1a2e",
                relief=tk.FLAT,
                padx=8,
                pady=6,
                cursor="hand2",
            ).pack(side=tk.LEFT)

        # Step 1 — prize list
        block1 = _step_block(
            1,
            "Prize list file",
            "Excel or tab-separated file with SKU, Qty, and optional img column "
            "(image file name or URL).",
        )
        self.path_var = tk.StringVar(value=str(draw_prize.default_list_path()))
        self.path_var.trace_add("write", self._on_path_write)
        self.path_entry = _path_entry_row(block1, self.path_var)
        _action_buttons(
            block1,
            pick_cmd=self._browse,
            pick_label="  Choose prize list…  ",
            open_cmd=self._open_list_file_folder,
        )
        self._setup_list_status = tk.Label(
            block1,
            text="",
            font=("Segoe UI", 9),
            bg=card_bg,
            fg=WHEEL_MUTED,
            anchor=tk.W,
            justify=tk.LEFT,
            wraplength=400,
        )
        self._setup_list_status.pack(fill=tk.X, pady=(8, 0))
        self._setup_wrap_labels.append(self._setup_list_status)

        # Step 2 — images folder
        default_images = script_dir() / "Images"
        if not default_images.is_dir():
            default_images = script_dir()

        block2 = _step_block(
            2,
            "Images folder",
            "All pictures for the wheel live here. In the list, img can be just the file name "
            '(e.g. GrassEnergy.jpg) or a subpath like images/GrassEnergy.jpg.',
        )
        self.images_path_var = tk.StringVar(value=str(default_images))
        self.images_path_var.trace_add("write", self._on_images_path_write)
        self.images_path_entry = _path_entry_row(block2, self.images_path_var)
        _action_buttons(
            block2,
            pick_cmd=self._browse_images,
            pick_label="  Choose images folder…  ",
            open_cmd=self._open_images_folder,
        )
        self._setup_images_status = tk.Label(
            block2,
            text="",
            font=("Segoe UI", 9),
            bg=card_bg,
            fg=WHEEL_MUTED,
            anchor=tk.W,
            justify=tk.LEFT,
            wraplength=400,
        )
        self._setup_images_status.pack(fill=tk.X, pady=(8, 0))
        self._setup_wrap_labels.append(self._setup_images_status)

        self._build_wheel_preset_panel(
            pad,
            bg=card_bg,
            card_bg=card_bg,
            border=card_edge,
            wraplength=400,
        )

        quick = tk.Frame(pad, bg=card_bg)
        quick.pack(fill=tk.X, pady=(0, 4))
        tk.Button(
            quick,
            text="  Reset saved session…  ",
            command=self._setup_reset_persisted_settings,
            font=("Segoe UI", 9),
            bg=BTN_KEEP,
            fg="white",
            activebackground="#a93226",
            activeforeground="white",
            relief=tk.FLAT,
            padx=8,
            pady=8,
            cursor="hand2",
        ).pack(anchor=tk.W)

        self._setup_row2 = tk.Frame(parent, bg=self._void_bg())
        self._setup_row2.pack(fill=tk.X, padx=4, pady=(0, 6))
        self._register_void_bg(self._setup_row2)
        self.dry_run_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(self._setup_row2, text="Dry run (never write file)", variable=self.dry_run_var).pack(
            side=tk.LEFT, padx=(0, 16)
        )
        ttk.Checkbutton(self._setup_row2, text="Show event log", variable=self._log_visible, command=self._toggle_log).pack(
            side=tk.LEFT
        )

        self.log_frame = ttk.Frame(parent)
        ttk.Label(self.log_frame, text="Log:").pack(anchor=tk.W)
        self.out = scrolledtext.ScrolledText(self.log_frame, height=6, wrap=tk.WORD, font=("Consolas", 9))
        self.out.pack(fill=tk.BOTH, expand=True, pady=(2, 0))

        self.after_idle(self._refresh_setup_status)
        self.after_idle(self._setup_sync_scroll_region)

    def _build_wheel_preset_panel(
        self,
        parent: tk.Frame,
        *,
        bg: str,
        card_bg: str,
        border: str,
        wraplength: int,
    ) -> None:
        """Wheel preset id / match-from-file — Show setup and Spin & controls."""
        outer = tk.Frame(
            parent,
            bg=card_bg,
            highlightbackground=border,
            highlightthickness=1,
        )
        outer.pack(fill=tk.X, pady=(0, 10))
        pad = tk.Frame(outer, bg=card_bg)
        pad.pack(fill=tk.X, padx=10, pady=10)
        tk.Label(
            pad,
            text="Wheel preset",
            font=("Segoe UI", 9, "bold"),
            bg=card_bg,
            fg=WHEEL_POINTER,
            anchor=tk.W,
        ).pack(fill=tk.X)
        tk.Label(
            pad,
            text="Type an id (e.g. wheel58912) or Match from file…. Paths are saved in energy_break_state.json. "
            "Switching presets reuses an existing winner spreadsheet for that wheel when one is in winner_sessions/.",
            font=("Segoe UI", 8),
            bg=card_bg,
            fg=WHEEL_MUTED,
            anchor=tk.W,
            justify=tk.LEFT,
            wraplength=wraplength,
        ).pack(fill=tk.X, pady=(2, 6))
        preset_row = tk.Frame(pad, bg=card_bg)
        preset_row.pack(fill=tk.X)
        preset_entry = tk.Entry(
            preset_row,
            textvariable=self.wheel_preset_var,
            font=("Segoe UI", 9),
            bg=INV_SLOT_FACE,
            fg=WHEEL_FG,
            insertbackground=WHEEL_FG,
            highlightthickness=1,
            highlightbackground=WHEEL_CELL_BG,
            highlightcolor=WHEEL_POINTER,
        )
        preset_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        preset_entry.bind("<Return>", lambda _e: self._setup_apply_wheel_preset())
        actions = tk.Frame(pad, bg=card_bg)
        actions.pack(fill=tk.X, pady=(6, 0))
        tk.Button(
            actions,
            text="  Apply preset  ",
            command=self._setup_apply_wheel_preset,
            font=("Segoe UI", 9, "bold"),
            bg=BTN_SPIN,
            fg="white",
            activebackground="#3498db",
            activeforeground="white",
            relief=tk.FLAT,
            padx=8,
            pady=6,
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(
            actions,
            text="  Match from file…  ",
            command=self._setup_match_wheel_preset_from_file,
            font=("Segoe UI", 9),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            activebackground=WHEEL_POINTER,
            activeforeground="#1a1a2e",
            relief=tk.FLAT,
            padx=8,
            pady=6,
            cursor="hand2",
        ).pack(side=tk.LEFT)

    def _build_spin_controls_inner(self, ctrl_inner: tk.Frame) -> None:
        """Populate the Spin & controls window (also used when recreating that Toplevel)."""
        panel = PRIZE_BOARD_CONTENT_BG
        card_bg = SPIN_CONTROLS_CARD_BG
        edge = SPIN_CONTROLS_CARD_BORDER
        ctrl_inner.configure(bg=panel)

        def _section_title(text: str) -> None:
            tk.Label(
                ctrl_inner,
                text=text,
                font=("Segoe UI", 10, "bold"),
                bg=panel,
                fg=WHEEL_POINTER,
                anchor=tk.W,
            ).pack(fill=tk.X, pady=(4, 8))

        self._build_wheel_preset_panel(
            ctrl_inner,
            bg=panel,
            card_bg=card_bg,
            border=edge,
            wraplength=640,
        )
        tk.Label(
            ctrl_inner,
            text="Strip wheel + status below; HTML/OBS wheel uses the same live feed.",
            font=("Segoe UI", 9),
            bg=panel,
            fg=WHEEL_MUTED,
            wraplength=640,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 8))

        _section_title("Live wheel")
        card_live = tk.Frame(ctrl_inner, bg=card_bg, highlightbackground=edge, highlightthickness=1)
        card_live.pack(fill=tk.X, pady=(0, 14))
        live_pad = tk.Frame(card_live, bg=card_bg)
        live_pad.pack(fill=tk.X, padx=14, pady=14)
        live_chrome = tk.Frame(live_pad, bg=card_bg)
        live_chrome.pack(fill=tk.X, pady=(0, 8))
        self._sc_spin_counter_label = tk.Label(
            live_chrome,
            text="Spot 1",
            font=("Segoe UI", 12, "bold"),
            bg=card_bg,
            fg=WHEEL_POINTER,
            anchor=tk.W,
        )
        self._sc_spin_counter_label.pack(side=tk.LEFT)
        self._sc_prizes_label = tk.Label(
            live_chrome,
            text="Total prizes: —",
            font=("Segoe UI", 10, "bold"),
            bg=card_bg,
            fg=WHEEL_POINTER,
            anchor=tk.E,
        )
        self._sc_prizes_label.pack(side=tk.RIGHT)
        self.spin_controls_wheel_canvas = tk.Canvas(
            live_pad,
            height=SPIN_CONTROLS_WHEEL_CANVAS_H,
            bg=WHEEL_BG,
            highlightthickness=0,
            bd=0,
        )
        self.spin_controls_wheel_canvas.pack(fill=tk.X, expand=False)
        self.spin_controls_wheel_canvas.bind("<Configure>", self._on_wheel_canvas_configure)
        self._sc_wheel_status = tk.Label(
            live_pad,
            text="",
            font=("Segoe UI", 9, "italic"),
            bg=card_bg,
            fg=WHEEL_FG,
            wraplength=640,
            justify=tk.CENTER,
        )
        self._sc_wheel_status.pack(fill=tk.X, pady=(8, 0))

        _section_title("Draw")
        card_draw = tk.Frame(ctrl_inner, bg=card_bg, highlightbackground=edge, highlightthickness=1)
        card_draw.pack(fill=tk.X, pady=(0, 14))
        draw_pad = tk.Frame(card_draw, bg=card_bg)
        draw_pad.pack(fill=tk.X, padx=14, pady=14)

        tk.Label(
            draw_pad,
            text=(
                "SPIN — Run the Wheel; the Prize is committed.\n"
                "SUPER SPIN — Run the Wheel; Allowed to Re-Roll Once. "
                "REROLL and KEEP appear in this row only after a Super spin stops."
            ),
            bg=card_bg,
            fg=WHEEL_FG,
            font=("Segoe UI", 9),
            wraplength=640,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 12))

        speed_row = tk.Frame(draw_pad, bg=card_bg)
        speed_row.pack(fill=tk.X, pady=(0, 10))
        tk.Label(
            speed_row,
            text="Spin speed",
            bg=card_bg,
            fg=WHEEL_FG,
            font=("Segoe UI", 9, "bold"),
        ).pack(side=tk.LEFT, padx=(0, 12))
        self.wheel_spin_speed_var = tk.StringVar(value="fast")
        for label, val in (("Fast (~2 s)", "fast"), ("Slow (~4.5 s)", "slow")):
            tk.Radiobutton(
                speed_row,
                text=label,
                variable=self.wheel_spin_speed_var,
                value=val,
                bg=card_bg,
                fg=WHEEL_FG,
                activebackground=card_bg,
                activeforeground=WHEEL_POINTER,
                selectcolor=WHEEL_CELL_BG,
                font=("Segoe UI", 9),
                highlightthickness=0,
                cursor="hand2",
            ).pack(side=tk.LEFT, padx=(0, 14))

        self.control_bar = tk.Frame(draw_pad, bg=card_bg)
        self.control_bar.pack(fill=tk.X)

        btn_kw: dict[str, object] = {
            "font": self._btn_font,
            "relief": tk.FLAT,
            "bd": 0,
            "padx": 16,
            "pady": 12,
            "cursor": "hand2",
        }
        self.spin_btn = tk.Button(
            self.control_bar,
            text="  SPIN  ",
            command=self._on_spin,
            bg=BTN_SPIN,
            fg="white",
            activebackground="#3498db",
            activeforeground="white",
            **btn_kw,
        )
        self.spin_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.super_btn = tk.Button(
            self.control_bar,
            text="  SUPER SPIN  ",
            command=self._on_super_spin,
            bg=BTN_SUPER,
            fg="white",
            activebackground="#9b59b6",
            activeforeground="white",
            **btn_kw,
        )
        self.super_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.reroll_btn = tk.Button(
            self.control_bar,
            text="  REROLL  ",
            command=self._on_reroll,
            bg=BTN_REROLL,
            fg="white",
            activebackground="#1abc9c",
            activeforeground="white",
            **btn_kw,
        )
        self.keep_btn = tk.Button(
            self.control_bar,
            text="  KEEP  ",
            command=self._on_keep,
            bg=BTN_KEEP,
            fg="white",
            activebackground="#e74c3c",
            activeforeground="white",
            **btn_kw,
        )

        _section_title("Edit spin")
        card_edit = tk.Frame(ctrl_inner, bg=card_bg, highlightbackground=edge, highlightthickness=1)
        card_edit.pack(fill=tk.X, pady=(0, 14))
        edit_pad = tk.Frame(card_edit, bg=card_bg)
        edit_pad.pack(fill=tk.X, padx=14, pady=14)
        tk.Label(
            edit_pad,
            text=(
                "Skip spot — Skip a # in the Saved Sequence (use when payment fails or pending).\n"
                "Undo spin — If you accidentally spun, press this button."
            ),
            bg=card_bg,
            fg=WHEEL_FG,
            font=("Segoe UI", 9),
            wraplength=640,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 12))
        edit_btn_row = tk.Frame(edit_pad, bg=card_bg)
        edit_btn_row.pack(fill=tk.X)
        self.skip_spot_btn = tk.Button(
            edit_btn_row,
            text="  Skip spot  ",
            command=self._on_skip_spot,
            font=("Segoe UI", 10, "bold"),
            bg=BTN_SKIP,
            fg="white",
            activebackground="#eb984e",
            activeforeground="white",
            relief=tk.FLAT,
            bd=0,
            padx=14,
            pady=10,
            cursor="hand2",
        )
        self.skip_spot_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.undo_spin_btn = tk.Button(
            edit_btn_row,
            text="  Undo spin  ",
            command=self._on_undo_spin,
            font=("Segoe UI", 10, "bold"),
            bg=BTN_UNDO,
            fg="white",
            activebackground="#2874a6",
            activeforeground="white",
            state=tk.DISABLED,
            relief=tk.FLAT,
            bd=0,
            padx=14,
            pady=10,
            cursor="hand2",
        )
        self.undo_spin_btn.pack(side=tk.LEFT, padx=(0, 10))

        _section_title("Fill skipped spot")
        card_fill = tk.Frame(ctrl_inner, bg=card_bg, highlightbackground=edge, highlightthickness=1)
        card_fill.pack(fill=tk.X, pady=(0, 14))
        fill_pad = tk.Frame(card_fill, bg=card_bg)
        fill_pad.pack(fill=tk.X, padx=14, pady=14)
        tk.Label(
            fill_pad,
            text=(
                "Skipped spot still empty? Click Fill a skipped spot…, pick the row in the list, then run one "
                "normal SPIN. Cancel fill mode to return to the usual next spot #."
            ),
            bg=card_bg,
            fg=WHEEL_MUTED,
            font=("Segoe UI", 9),
            wraplength=640,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 8))
        self._fill_skipped_hint_label = tk.Label(
            fill_pad,
            text="",
            bg=card_bg,
            fg=WHEEL_MUTED,
            font=("Segoe UI", 8),
            wraplength=640,
            justify=tk.LEFT,
            anchor=tk.W,
        )
        self._fill_skipped_hint_label.pack(fill=tk.X, pady=(0, 8))
        self._fill_skipped_status_label = tk.Label(
            fill_pad,
            text="",
            bg=card_bg,
            fg=WHEEL_POINTER,
            font=("Segoe UI", 9),
            wraplength=640,
            justify=tk.LEFT,
            anchor=tk.W,
        )
        self._fill_skipped_status_label.pack(fill=tk.X, pady=(0, 8))
        fill_btns = tk.Frame(fill_pad, bg=card_bg)
        fill_btns.pack(fill=tk.X)
        tk.Button(
            fill_btns,
            text="  Fill a skipped spot…  ",
            command=self._open_fill_skipped_spot_dialog,
            font=("Segoe UI", 10, "bold"),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            activebackground=WHEEL_POINTER,
            activeforeground="#1a1a2e",
            relief=tk.FLAT,
            bd=0,
            padx=12,
            pady=8,
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(
            fill_btns,
            text="  Cancel fill mode  ",
            command=self.cancel_fill_skipped_spot_mode,
            font=("Segoe UI", 10),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            activebackground=WHEEL_POINTER,
            activeforeground="#1a1a2e",
            relief=tk.FLAT,
            bd=0,
            padx=12,
            pady=8,
            cursor="hand2",
        ).pack(side=tk.LEFT)

        _section_title("HTML wheel (OBS)")
        card_html = tk.Frame(ctrl_inner, bg=card_bg, highlightbackground=edge, highlightthickness=1)
        card_html.pack(fill=tk.X, pady=(0, 14))
        html_pad = tk.Frame(card_html, bg=card_bg)
        html_pad.pack(fill=tk.X, padx=14, pady=12)
        tk.Label(
            html_pad,
            text="Browser wheel for OBS (localhost). Use the same URL in a Browser Source after the app prints the port.",
            bg=card_bg,
            fg=WHEEL_MUTED,
            font=("Segoe UI", 9),
            wraplength=520,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 10))
        html_row = tk.Frame(html_pad, bg=card_bg)
        html_row.pack(fill=tk.X)
        tk.Button(
            html_row,
            text="  Open HTML wheel…  ",
            command=self._open_html_wheel_in_browser,
            font=("Segoe UI", 10, "bold"),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            activebackground=WHEEL_POINTER,
            activeforeground="#1a1a2e",
            relief=tk.FLAT,
            bd=0,
            padx=14,
            pady=10,
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=(0, 10))

        _section_title("Prize board")
        card_pb = tk.Frame(ctrl_inner, bg=card_bg, highlightbackground=edge, highlightthickness=1)
        card_pb.pack(fill=tk.X, pady=(0, 14))
        pb_pad = tk.Frame(card_pb, bg=card_bg)
        pb_pad.pack(fill=tk.X, padx=14, pady=12)
        pb_row = tk.Frame(pb_pad, bg=card_bg)
        pb_row.pack(fill=tk.X)
        tk.Label(
            pb_row,
            text="Board:",
            bg=card_bg,
            fg=WHEEL_FG,
            font=("Segoe UI", 9, "bold"),
        ).pack(side=tk.LEFT, padx=(0, 6))
        tk.Label(pb_row, text="columns", bg=card_bg, fg=WHEEL_MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Spinbox(
            pb_row,
            from_=2,
            to=24,
            width=4,
            textvariable=self._prize_board_grid_cols,
            font=("Segoe UI", 10),
            bg=INV_SLOT_FACE,
            fg=WHEEL_FG,
            buttonbackground=WHEEL_CELL_BG,
            highlightthickness=0,
        ).pack(side=tk.LEFT, padx=(4, 10))
        tk.Label(pb_row, text="×", bg=card_bg, fg=WHEEL_MUTED, font=("Segoe UI", 10)).pack(side=tk.LEFT)
        tk.Spinbox(
            pb_row,
            from_=2,
            to=20,
            width=4,
            textvariable=self._prize_board_tile_rows_fit,
            font=("Segoe UI", 10),
            bg=INV_SLOT_FACE,
            fg=WHEEL_FG,
            buttonbackground=WHEEL_CELL_BG,
            highlightthickness=0,
        ).pack(side=tk.LEFT, padx=(4, 6))
        tk.Label(
            pb_row,
            text="rows (tile size)",
            bg=card_bg,
            fg=WHEEL_MUTED,
            font=("Segoe UI", 9),
        ).pack(side=tk.LEFT)
        tk.Label(
            pb_pad,
            text="Applies to the separate Prize board window (columns per row; rows set how tile height fits the panel).",
            bg=card_bg,
            fg=WHEEL_MUTED,
            font=("Segoe UI", 9),
            wraplength=520,
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(10, 0))

        self.btn_row = self.control_bar

    def _clear_undo_spin(self) -> None:
        self._undo_spin_snap = None
        self._sync_undo_spin_button()

    def _sync_undo_spin_button(self) -> None:
        if not hasattr(self, "undo_spin_btn"):
            return
        can = self._undo_spin_snap is not None and not self._busy
        if can:
            self.undo_spin_btn.config(state=tk.NORMAL, bg=BTN_UNDO)
        else:
            self.undo_spin_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)

    def _push_undo_spin(
        self,
        list_path: Path,
        result: draw_prize.SpinResult,
        spot_before_append: int,
        winner_logged: bool,
    ) -> None:
        self._undo_spin_snap = _UndoSpinSnapshot(
            "spin",
            list_path.resolve(),
            result,
            spot_before_append,
            winner_logged,
        )
        self._sync_undo_spin_button()

    def _push_undo_skip_log(self, spot_written: int) -> None:
        try:
            lp = self._list_path().resolve()
        except OSError:
            lp = Path()
        self._undo_spin_snap = _UndoSpinSnapshot(
            "skip_log",
            lp,
            None,
            spot_written,
            True,
        )
        self._sync_undo_spin_button()

    def _push_undo_skip_counter_only(self, spot_before: int) -> None:
        try:
            lp = self._list_path().resolve()
        except OSError:
            lp = Path()
        self._undo_spin_snap = _UndoSpinSnapshot(
            "skip_counter_only",
            lp,
            None,
            spot_before,
            False,
        )
        self._sync_undo_spin_button()

    def _push_undo_backfill(
        self,
        list_path: Path,
        result: draw_prize.SpinResult,
        spot_filled: int,
        sheet_updated: bool,
    ) -> None:
        self._undo_spin_snap = _UndoSpinSnapshot(
            "backfill",
            list_path.resolve(),
            result,
            spot_filled,
            sheet_updated,
        )
        self._sync_undo_spin_button()

    def _winner_delete_row_for_spot(self, path: Path, spot_number: int) -> bool:
        """Remove the data row whose column A equals ``spot_number`` (last match if duplicated)."""
        if load_workbook is None or Workbook is None:
            return False
        try:
            wb = load_workbook(path)
            ws = wb.active
            if ws is None:
                wb.close()
                return False
            delete_at: int | None = None
            max_r = ws.max_row or 1
            for r in range(2, max_r + 1):
                a = ws.cell(row=r, column=1).value
                try:
                    n = int(a) if a is not None and str(a).strip() != "" else None
                except (TypeError, ValueError):
                    n = None
                if n == spot_number:
                    delete_at = r
            if delete_at is None:
                wb.close()
                return False
            ws.delete_rows(delete_at, 1)
            wb.save(path)
            wb.close()
            return True
        except Exception:
            return False

    def _apply_undo_spin_snapshot(self, snap: _UndoSpinSnapshot) -> str | None:
        """Return an error message, or None on success."""
        try:
            if snap.kind == "skip_counter_only":
                self._winner_next_spin = snap.spot_written
                return None
            if snap.kind == "skip_log":
                wp = getattr(self, "_winner_log_path", None)
                if wp and snap.winner_row_added:
                    if not self._winner_delete_row_for_spot(Path(wp), snap.spot_written):
                        return "Could not remove the skip row from the winner workbook (close Excel if it is open)."
                self._winner_next_spin = snap.spot_written
                return None
            if snap.kind == "backfill":
                if snap.result is None:
                    return "Internal error: bad undo snapshot."
                if snap.list_path.resolve() != Path(self._list_path()).resolve():
                    return "The prize list path changed — undo aborted."
                sess = draw_prize.open_draw_session(snap.list_path)
                sess.revert_decrement(snap.result)
                if snap.winner_row_added:
                    wp = getattr(self, "_winner_log_path", None)
                    if wp and not self._winner_clear_prize_cell_for_spot(Path(wp), snap.spot_written):
                        return (
                            "Prize quantity was restored, but the winner sheet cell could not be cleared. "
                            "Clear column B for that spot manually if needed."
                        )
                if hasattr(self, "_winner_spot_lookup_cache"):
                    self._winner_spot_lookup_cache = None
                return None
            if snap.kind != "spin" or snap.result is None:
                return "Internal error: bad undo snapshot."
            if snap.list_path.resolve() != Path(self._list_path()).resolve():
                return "The prize list path changed — undo aborted."
            sess = draw_prize.open_draw_session(snap.list_path)
            sess.revert_decrement(snap.result)
            if snap.winner_row_added:
                wp = getattr(self, "_winner_log_path", None)
                if wp:
                    if not self._winner_delete_row_for_spot(Path(wp), snap.spot_written):
                        return (
                            "Prize quantity was restored, but the winner sheet row could not be removed. "
                            "Delete that row manually if needed."
                        )
            self._winner_next_spin = snap.spot_written
            return None
        except draw_prize.PrizeDrawError as e:
            return str(e)
        except Exception as e:
            return f"{type(e).__name__}: {e}"

    def _set_overlay_winner_list_height(self) -> None:
        """Cap listbox viewport to OVERLAY_WINNER_LIST_VISIBLE_ROWS; scroll when there are more rows."""
        if not hasattr(self, "_overlay_listbox"):
            return
        lb = self._overlay_listbox
        n = max(1, lb.size())
        cap = int(type(self).OVERLAY_WINNER_LIST_VISIBLE_ROWS)
        lb.configure(height=min(cap, n))

    def _on_undo_spin(self) -> None:
        if self._busy:
            return
        snap = self._undo_spin_snap
        if snap is None:
            return
        if not messagebox.askyesno(
            "Undo spin",
            "Reverse the last completed spin or skip?\n\n"
            "This restores the prize list quantity if it changed, removes the matching winner row if one was logged, "
            "and steps the spot counter back. Super / Reroll in progress will be cancelled.",
        ):
            return
        err = self._apply_undo_spin_snapshot(snap)
        if err:
            messagebox.showerror("Undo spin", err)
            self._log(f"Undo failed: {err}\n")
            return
        self._hide_super_panel()
        self._clear_undo_spin()
        if hasattr(self, "_winner_spot_lookup_cache"):
            self._winner_spot_lookup_cache = None
        self._cancel_pulse()
        self._invalidate_session()
        self._update_spin_counter_label()
        self._reset_wheel_idle()
        self._refresh_prizes_label()
        self._update_draw_buttons_for_supply()
        self._wheel_publish_html_snapshot()
        self._log("Undo: last draw / skip reversed.\n")

    def _build_wheel_area(self, parent: tk.Frame) -> None:
        """HTML wheel is primary for OBS; Tk canvas lives off-screen for spin state and /api snapshots."""
        parent.configure(bg=self._void_bg())
        vb = self._void_bg()

        compact = tk.Frame(parent, bg=vb)
        compact.pack(fill=tk.X, expand=False, padx=12, pady=(12, 10))
        self._register_void_bg(compact)
        tk.Label(
            compact,
            text="Strip wheel + status run in Spin & controls and in the HTML page (OBS). "
            "Wheel presets are under Show setup.",
            font=("Segoe UI", 10),
            bg=vb,
            fg=WHEEL_FG,
            wraplength=460,
            justify=tk.CENTER,
        ).pack(fill=tk.X, pady=(0, 12))
        row = tk.Frame(compact, bg=vb)
        row.pack(fill=tk.X)
        self._register_void_bg(row)
        self._prize_board_btn = tk.Button(
            row,
            text="  Prize board  ",
            command=self._show_prize_board,
            font=("Segoe UI", 10, "bold"),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            activebackground=WHEEL_POINTER,
            activeforeground="#1a1a2e",
            relief=tk.FLAT,
            padx=12,
            pady=6,
        )
        self._prize_board_btn.pack(side=tk.LEFT, padx=(0, 10))
        self._spin_controls_title_btn = tk.Button(
            row,
            text="  Spin & controls…  ",
            command=self._show_spin_controls,
            font=("Segoe UI", 10, "bold"),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            activebackground=WHEEL_POINTER,
            activeforeground="#1a1a2e",
            relief=tk.FLAT,
            padx=12,
            pady=6,
        )
        self._spin_controls_title_btn.pack(side=tk.LEFT)

        winner_pick = tk.Frame(compact, bg=vb)
        winner_pick.pack(fill=tk.X, pady=(10, 0))
        self._winner_log_path_label = tk.Label(
            winner_pick,
            text="",
            bg=vb,
            fg=WHEEL_POINTER,
            font=("Segoe UI", 9, "bold"),
            wraplength=460,
            justify=tk.LEFT,
            anchor=tk.W,
        )
        self._winner_log_path_label.pack(fill=tk.X, pady=(0, 6))
        tk.Button(
            winner_pick,
            text="  Choose winner spreadsheet…  ",
            command=self._choose_winner_session_spreadsheet,
            font=("Segoe UI", 10, "bold"),
            bg=BTN_SPIN,
            fg="white",
            activebackground="#3498db",
            activeforeground="white",
            relief=tk.FLAT,
            bd=0,
            padx=12,
            pady=8,
            cursor="hand2",
        ).pack(anchor=tk.W)
        self.after_idle(self._refresh_winner_log_path_label)

        self._overlay_list_outer = tk.Frame(
            compact,
            bg=WHEEL_FRAME_BG,
            highlightbackground=WHEEL_CELL_BORDER,
            highlightthickness=1,
        )
        self._overlay_list_outer.pack(fill=tk.X, expand=False, pady=(10, 0))
        self._overlay_list_title = tk.Label(
            self._overlay_list_outer,
            text="Prize Wheel",
            bg=WHEEL_FRAME_BG,
            fg=WHEEL_TITLE,
            font=("Segoe UI", 10, "bold"),
            anchor=tk.W,
        )
        self._overlay_list_title.pack(fill=tk.X, padx=8, pady=(8, 4))
        self._overlay_spot_target_label = tk.Label(
            self._overlay_list_outer,
            text="",
            bg=WHEEL_FRAME_BG,
            fg=WHEEL_POINTER,
            font=("Segoe UI", 10, "bold"),
            anchor=tk.W,
        )
        self._overlay_spot_target_label.pack(fill=tk.X, padx=8, pady=(0, 4))
        list_row = tk.Frame(self._overlay_list_outer, bg=WHEEL_FRAME_BG)
        list_row.pack(fill=tk.X, padx=6, pady=(0, 8))
        self._overlay_list_scroll = tk.Scrollbar(
            list_row,
            orient=tk.VERTICAL,
            bg=WHEEL_FRAME_BG,
            troughcolor=WHEEL_CELL_BG,
        )
        self._overlay_listbox = tk.Listbox(
            list_row,
            height=1,
            font=("Segoe UI", 10),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            selectbackground=WHEEL_POINTER,
            selectforeground="#1a1a2e",
            highlightthickness=0,
            activestyle="dotbox",
            exportselection=False,
            yscrollcommand=self._overlay_list_scroll.set,
        )
        self._overlay_list_scroll.config(command=self._overlay_listbox.yview)
        self._overlay_list_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self._overlay_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        shw, shh = 1000, int(round(280 * WHEEL_VERTICAL_SCALE))
        self._wheel_sensor_host = tk.Frame(self, bg=vb, width=shw, height=shh)
        self._wheel_sensor_host.place(x=-3200, y=0)
        self._wheel_sensor_host.pack_propagate(False)
        self._register_void_bg(self._wheel_sensor_host)

        self._wheel_heading_block = tk.Frame(self._wheel_sensor_host, bg=vb)
        self._wheel_heading_block.pack(fill=tk.X, padx=10, pady=(8, 2))
        self._wheel_heading_block.grid_columnconfigure(0, weight=1)
        self._wheel_heading_block.grid_columnconfigure(1, weight=0)
        self._wheel_heading_block.grid_columnconfigure(2, weight=1)
        self._register_void_bg(self._wheel_heading_block)

        self._wheel_heading_label = tk.Label(
            self._wheel_heading_block,
            text="Prize Wheel",
            font=("Segoe UI", 20, "bold"),
            bg=vb,
            fg=WHEEL_TITLE,
            anchor=tk.CENTER,
        )
        self._wheel_heading_label.grid(row=0, column=1, sticky=tk.N)
        self._register_void_bg(self._wheel_heading_label)

        self.prizes_label = tk.Label(
            self._wheel_heading_block,
            text="Total prizes: —",
            font=("Segoe UI", 17, "bold"),
            bg=vb,
            fg=WHEEL_POINTER,
            anchor=tk.E,
            justify=tk.RIGHT,
        )
        self.prizes_label.grid(row=0, column=2, sticky=tk.NE)
        self._register_void_bg(self.prizes_label)

        self.spin_counter_label = tk.Label(
            self._wheel_sensor_host,
            text="Spot 1",
            font=("Segoe UI", 20, "bold"),
            bg=vb,
            fg=WHEEL_POINTER,
        )
        self.spin_counter_label.pack(fill=tk.X, padx=10, pady=(0, 2))
        self._register_void_bg(self.spin_counter_label)

        self.wheel_canvas = tk.Canvas(
            self._wheel_sensor_host,
            height=self.WHEEL_CELL,
            bg=vb,
            highlightthickness=0,
        )
        self.wheel_canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=(2, 4))
        self._register_void_bg(self.wheel_canvas)
        self.wheel_canvas.bind("<Configure>", self._on_wheel_canvas_configure)

        self.wheel_status = tk.Label(
            self._wheel_sensor_host,
            text="",
            font=("Segoe UI", 10, "italic"),
            bg=vb,
            fg=WHEEL_FG,
            wraplength=shw - 32,
            justify=tk.CENTER,
        )
        self.wheel_status.pack(pady=(0, 6))
        self._register_void_bg(self.wheel_status)

    def _wheel_display_canvases(self) -> list[tk.Canvas]:
        out: list[tk.Canvas] = []
        if hasattr(self, "wheel_canvas"):
            out.append(self.wheel_canvas)
        sc = getattr(self, "spin_controls_wheel_canvas", None)
        if sc is not None:
            try:
                if sc.winfo_exists():
                    out.append(sc)
            except tk.TclError:
                pass
        return out

    def _on_spin_controls_wheel_panel_ready(self) -> None:
        """Refresh live wheel + chrome when Spin & controls opens."""
        sc = getattr(self, "spin_controls_wheel_canvas", None)
        if sc is not None:
            try:
                if sc.winfo_exists():
                    sc.configure(height=SPIN_CONTROLS_WHEEL_CANVAS_H)
            except tk.TclError:
                pass
        self._mirror_wheel_chrome()
        self._schedule_spin_controls_wheel_sync()
        self._refresh_winner_log_path_label()
        if hasattr(self, "wheel_canvas"):
            self._wheel_redraw()

    def _mirror_wheel_chrome(self) -> None:
        if hasattr(self, "spin_counter_label") and hasattr(self, "_sc_spin_counter_label"):
            try:
                self._sc_spin_counter_label.configure(text=str(self.spin_counter_label.cget("text")))
            except tk.TclError:
                pass
        if hasattr(self, "prizes_label") and hasattr(self, "_sc_prizes_label"):
            try:
                self._sc_prizes_label.configure(text=str(self.prizes_label.cget("text")))
            except tk.TclError:
                pass
        if hasattr(self, "wheel_status") and hasattr(self, "_sc_wheel_status"):
            try:
                self._sc_wheel_status.configure(
                    text=str(self.wheel_status.cget("text")),
                    fg=str(self.wheel_status.cget("fg")),
                )
            except tk.TclError:
                pass

    def _set_wheel_status(self, text: str, fg: str = WHEEL_FG) -> None:
        if hasattr(self, "wheel_status"):
            try:
                self.wheel_status.configure(text=text, fg=fg)
            except tk.TclError:
                pass
        sc = getattr(self, "_sc_wheel_status", None)
        if sc is not None:
            try:
                if sc.winfo_exists():
                    sc.configure(text=text, fg=fg)
            except tk.TclError:
                pass

    def _on_wheel_canvas_configure(self, _event: tk.Event | None = None) -> None:
        """Keep the strip centered under the pointer when the primary (OBS) canvas width changes."""
        ref_w = self._wheel_reference_viewport_w()
        cw_px = self._wheel_primary_viewport_w()
        if abs(cw_px - ref_w) > 0.5:
            self._wheel_target_scroll = self._wheel_scroll_convert_viewport(
                float(self._wheel_target_scroll), ref_w, cw_px
            )
            self._wheel_scroll = self._wheel_scroll_convert_viewport(
                float(self._wheel_scroll), ref_w, cw_px
            )
            self._wheel_note_scroll_viewport(cw_px)
        if (
            not self._busy
            and self._anim_after is None
            and self._pulse_after is None
            and self._pending_super is None
            and len(self._wheel_strip)
            == int(type(self).WHEEL_IDLE_STRIP_LEN) * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES)
        ):
            mid = self._wheel_win_idx
            self._wheel_target_scroll = self._wheel_scroll_to_center_index(float(mid), cw_px)
            self._wheel_scroll = self._wheel_target_scroll
            self._wheel_idle_offset = 0.0
            self._wheel_note_scroll_viewport(cw_px)
        self._wheel_redraw()

    def _cancel_wheel_idle_drift(self) -> None:
        aid = getattr(self, "_wheel_idle_drift_after", None)
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
            self._wheel_idle_drift_after = None
        self._wheel_idle_offset = 0.0

    def _wheel_idle_drift_active(self) -> bool:
        if self._busy or self._anim_after is not None or self._pulse_after is not None:
            return False
        if self._pending_super is not None:
            return False
        if getattr(self, "_customer_empty_wheel", False):
            return False
        if (
            len(self._wheel_strip)
            != int(type(self).WHEEL_IDLE_STRIP_LEN) * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES)
        ):
            return False
        if self._wheel_banner_title and not self._wheel_strip:
            return False
        return True

    def _wheel_scroll_for_strip_render(self) -> float:
        s = float(self._wheel_scroll)
        if self._wheel_idle_drift_active():
            return s + float(getattr(self, "_wheel_idle_offset", 0.0) or 0.0)
        return s

    def _wheel_idle_drift_tick(self) -> None:
        self._wheel_idle_drift_after = None
        if not self._wheel_idle_drift_active():
            self._wheel_idle_offset = 0.0
            return
        period = float(type(self).WHEEL_IDLE_STRIP_LEN) * float(self.WHEEL_CELL)
        if period < 1.0:
            return
        ms = float(int(type(self).WHEEL_IDLE_DRIFT_MS))
        delta = float(type(self).WHEEL_IDLE_DRIFT_PX_PER_SEC) * (ms / 1000.0)
        self._wheel_idle_offset = (self._wheel_idle_offset + delta) % period
        self._wheel_redraw(_idle_drift_frame=True)
        if self._wheel_idle_drift_active():
            self._wheel_idle_drift_after = self.after(
                int(type(self).WHEEL_IDLE_DRIFT_MS),
                self._wheel_idle_drift_tick,
            )

    def _spin_controls_wheel_link_reset(self) -> None:
        self._cancel_spin_controls_wheel_sync()
        self._wheel_control_link_ref = None

    def _cancel_spin_controls_wheel_sync(self) -> None:
        aid = self._spin_controls_wheel_sync_after
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
            self._spin_controls_wheel_sync_after = None

    def _schedule_spin_controls_wheel_sync(self) -> None:
        self._cancel_spin_controls_wheel_sync()
        self._spin_controls_wheel_sync_after = self.after(75, self._flush_spin_controls_wheel_sync)

    def _flush_spin_controls_wheel_sync(self) -> None:
        self._spin_controls_wheel_sync_after = None
        cw = self._controls_win
        if cw is None or not hasattr(self, "wheel_canvas"):
            return
        try:
            if not cw.winfo_exists():
                return
        except tk.TclError:
            return
        try:
            cww = int(cw.winfo_width())
            ch = int(cw.winfo_height())
        except tk.TclError:
            return
        if cww < 400 or ch < 340:
            return
        h_min = int(type(self).WHEEL_CANVAS_H_LINK_MIN)
        h_max = int(type(self).WHEEL_CANVAS_H_LINK_MAX)
        ref = self._wheel_control_link_ref
        if ref is None:
            try:
                self.update_idletasks()
            except tk.TclError:
                pass
            try:
                cur_canvas_h = int(self.wheel_canvas.cget("height"))
            except tk.TclError:
                return
            self._wheel_control_link_ref = (ch, cur_canvas_h)
            return
        ref_ch, ref_canvas_h = ref
        new_canvas_h = ref_canvas_h + (ch - ref_ch)
        new_canvas_h = max(h_min, min(h_max, new_canvas_h))
        try:
            cur_h = int(self.wheel_canvas.cget("height"))
        except tk.TclError:
            return
        resized = False
        if new_canvas_h != cur_h:
            self._wheel_image_cache.clear()
            try:
                self.wheel_canvas.configure(height=new_canvas_h)
            except tk.TclError:
                return
            resized = True
            try:
                self._ensure_wheel_photos_for_strip(self._wheel_strip)
            except tk.TclError:
                pass
        self._sync_spin_controls_wheel_canvas_height(ch)
        if resized:
            self._on_wheel_canvas_configure()

    def _sync_spin_controls_wheel_canvas_height(self, controls_win_h: int | None = None) -> None:
        """Keep the Spin & controls live preview compact (not tied to hidden wheel / OBS canvas height)."""
        sc = getattr(self, "spin_controls_wheel_canvas", None)
        if sc is None:
            return
        try:
            if not sc.winfo_exists():
                return
        except tk.TclError:
            return
        base = SPIN_CONTROLS_WHEEL_CANVAS_H
        h_max = SPIN_CONTROLS_WHEEL_CANVAS_H_MAX
        h_min = int(type(self).WHEEL_CANVAS_H_LINK_MIN)
        if controls_win_h is not None and controls_win_h >= 340:
            ref = self._wheel_control_link_ref
            if ref is not None:
                ref_ch, _ref_canvas_h = ref
                # Slight growth when the window is tall, but capped well below the OBS wheel.
                base = max(h_min, min(h_max, base + int((controls_win_h - ref_ch) * 0.12)))
        try:
            cur = int(sc.cget("height"))
        except tk.TclError:
            cur = 0
        if cur != base:
            try:
                sc.configure(height=base)
                self._wheel_redraw()
            except tk.TclError:
                pass

    def _btn_set_busy(self, btn: tk.Button, busy: bool) -> None:
        if busy:
            btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        else:
            btn.config(state=tk.NORMAL)

    def _total_qty_remaining(self) -> int | None:
        try:
            p = self._list_path()
        except OSError:
            return None
        if not p.is_file():
            return None
        try:
            _, rows = draw_prize.load_rows(p)
            return sum(q for _, q, _ in rows)
        except (draw_prize.PrizeDrawError, OSError, ValueError, TypeError):
            return None

    def _setup_paths_spin_block_reason(self) -> str | None:
        """None when prize list and images folder are set and exist."""
        if not hasattr(self, "path_var"):
            return None
        list_raw = self.path_var.get().strip()
        if not list_raw:
            return "Set a prize list in Show setup or apply a wheel preset."
        try:
            list_p = Path(list_raw).expanduser().resolve()
        except OSError:
            return "Prize list path is invalid."
        if not list_p.is_file():
            return f"Prize list not found ({list_p.name})."
        if not hasattr(self, "images_path_var"):
            return None
        images_raw = self.images_path_var.get().strip()
        if not images_raw:
            return "Set an images folder in Show setup or apply a wheel preset."
        try:
            images_p = Path(images_raw).expanduser().resolve()
        except OSError:
            return "Images folder path is invalid."
        if not images_p.is_dir():
            return f"Images folder not found ({images_p.name})."
        return None

    def _winner_session_spin_block_reason(self) -> str | None:
        """None when SPIN / Super / Skip may write to the active winner session workbook."""
        paths_block = self._setup_paths_spin_block_reason()
        if paths_block is not None:
            return paths_block
        if Workbook is None or load_workbook is None:
            return (
                "A winner session Excel file is required before spinning. "
                "Install openpyxl, then apply a wheel preset or choose a winner spreadsheet on the main tab."
            )
        path = getattr(self, "_winner_log_path", None)
        if path is None:
            return (
                "No winner session file is ready. Apply a wheel preset or choose a spreadsheet "
                "so each spin can be saved."
            )
        p = Path(path)
        if not p.is_file():
            return (
                f"Winner session file is missing ({p.name}). "
                "Apply the wheel preset again or choose a spreadsheet."
            )
        return None

    def _winner_session_ready_for_spin(self) -> bool:
        return self._winner_session_spin_block_reason() is None

    def _winnable_units_total(self) -> int | None:
        """Units still in the draw (sum of positive quantities). None if the list file cannot be read."""
        try:
            p = self._list_path()
        except OSError:
            return None
        if not p.is_file():
            return None
        try:
            _, rows = draw_prize.load_rows(p)
            return sum(q for _, q, _ in rows if q > 0)
        except (draw_prize.PrizeDrawError, OSError, ValueError, TypeError):
            return None

    def _style_buttons_depleted(self) -> None:
        if not hasattr(self, "spin_btn"):
            return
        self._hide_super_reroll_keep_buttons()
        self.spin_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self.super_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self.skip_spot_btn.config(state=tk.NORMAL, bg=BTN_SKIP)
        self._sync_undo_spin_button()

    def _wheel_paint_depleted_customer(self) -> None:
        """Clear the strip wheel and show a friendly message when nothing is left to draw."""
        self._cancel_wheel_idle_drift()
        self._wheel_sku_to_img = {}
        self._wheel_strip = []
        self._wheel_banner_title = "All prizes have been given out"
        self._wheel_banner_subtitle = (
            "Add stock to your prize list file, then save — new draws will unlock automatically."
        )
        self._wheel_banner_is_error = False
        self._wheel_paint_banner_on_canvases(
            "All prizes have been given out",
            "Add stock to your prize list file, then save — new draws will unlock automatically.",
            is_error=False,
            title_y_offset=-10,
            subtitle_y_offset=22,
        )
        self._set_wheel_status(
            "There are no prizes left on the wheel. Reload your list with more quantities to keep going.",
            WHEEL_TITLE,
        )
        self._wheel_publish_html_snapshot()

    def _paint_normal_idle_wheel_only(self) -> None:
        """Idle home strip: show real winnable SKUs (with images when available), not placeholder dashes."""
        self._wheel_banner_title = None
        self._wheel_banner_subtitle = None
        self._wheel_banner_is_error = False
        n = self.WHEEL_IDLE_STRIP_LEN
        mid = n // 2
        cw_px = max(int(self.wheel_canvas.winfo_width()), 520)
        self._wheel_win_idx = mid
        self._wheel_target_scroll = self._wheel_scroll_to_center_index(float(mid), float(cw_px))
        self._wheel_scroll = self._wheel_target_scroll
        self._wheel_note_scroll_viewport(cw_px)

        def _fallback_strip() -> None:
            self._wheel_sku_to_img = {}
            ln = n * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES)
            self._wheel_strip = ["···"] * ln

        try:
            p = self._list_path()
        except OSError:
            _fallback_strip()
            self._set_wheel_status(
                "SPIN = save now · SUPER SPIN = Reroll/Keep buttons appear after the wheel stops (no save until Keep)",
                WHEEL_FG,
            )
            self.after_idle(self._wheel_redraw)
            return
        if not p.is_file():
            _fallback_strip()
            self._set_wheel_status(
                "SPIN = save now · SUPER SPIN = Reroll/Keep buttons appear after the wheel stops (no save until Keep)",
                WHEEL_FG,
            )
            self.after_idle(self._wheel_redraw)
            return
        try:
            _, rows = draw_prize.load_rows(p)
        except draw_prize.PrizeDrawError:
            _fallback_strip()
            self._set_wheel_status(
                "SPIN = save now · SUPER SPIN = Reroll/Keep buttons appear after the wheel stops (no save until Keep)",
                WHEEL_FG,
            )
            self.after_idle(self._wheel_redraw)
            return

        self._wheel_sku_to_img = self._sku_to_img_map_from_rows(rows)
        first = draw_prize.sample_wheel_strip_labels(rows, n)
        if not first:
            self._wheel_strip = ["···"] * (n * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES))
            self._wheel_sku_to_img = {}
        else:
            self._wheel_strip = first + first
            self._ensure_wheel_photos_for_strip(self._wheel_strip)

        self._set_wheel_status(
            "SPIN = save now · SUPER SPIN = Reroll/Keep buttons appear after the wheel stops (no save until Keep)",
            WHEEL_FG,
        )
        self.after_idle(self._wheel_redraw)

    def _update_draw_buttons_for_supply(self) -> None:
        if not hasattr(self, "spin_btn") or self._busy:
            return
        units = self._winnable_units_total()
        depleted = units is not None and units == 0
        if depleted:
            self._style_buttons_depleted()
        elif not self._winner_session_ready_for_spin():
            self._style_buttons_no_winner_log()
        elif self._pending_super is not None:
            self._style_super_controls_active()
        else:
            self._style_main_controls_idle()

    def _style_buttons_no_winner_log(self) -> None:
        """Draw actions disabled until a winner session workbook exists on disk."""
        if not hasattr(self, "spin_btn"):
            return
        self._hide_super_reroll_keep_buttons()
        self.spin_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self.super_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self.skip_spot_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self._sync_undo_spin_button()

    def _reconcile_depleted_wheel_if_applicable(self) -> None:
        if not hasattr(self, "wheel_canvas"):
            return
        if self._busy or self._pending_super is not None or self._pulse_after is not None:
            return
        units = self._winnable_units_total()
        depleted = units is not None and units == 0
        if depleted:
            self._wheel_paint_depleted_customer()
            self._customer_empty_wheel = True
        elif self._customer_empty_wheel:
            self._customer_empty_wheel = False
            self._paint_normal_idle_wheel_only()

    def _update_prize_board_totals_label(
        self,
        rows_data: list[tuple[str, int, str]] | None,
        msg: str | None,
    ) -> None:
        lab = getattr(self, "_prize_board_totals_label", None)
        if lab is None:
            return
        try:
            if not lab.winfo_exists():
                return
        except tk.TclError:
            return
        if msg is not None:
            lab.configure(text="—")
            return
        if not rows_data:
            lab.configure(text="0 items available")
            return
        total_units = sum(q for _, q, _ in rows_data)
        n_types = len(rows_data)
        p = "prize types" if n_types != 1 else "prize type"
        lab.configure(text=f"{total_units} items available · {n_types} {p}")

    def _refresh_remaining_skus_panel(self) -> None:
        ri = getattr(self, "remaining_inner", None)
        if ri is None:
            self._update_prize_board_totals_label(None, "—")
            return
        try:
            if not ri.winfo_exists():
                return
        except tk.TclError:
            return
        for w in ri.winfo_children():
            w.destroy()
        msg: str | None = None
        rows_data: list[tuple[str, int, str]] = []
        try:
            p = self._list_path()
        except OSError:
            msg = "—"
        else:
            if not p.is_file():
                msg = "List file not found."
            else:
                try:
                    _, rows = draw_prize.load_rows(p)
                except draw_prize.PrizeDrawError as e:
                    msg = str(e)
                else:
                    rows_data = [(sku.strip(), q, img) for sku, q, img in rows if q > 0]
                    rows_data.sort(key=lambda t: t[0].lower())
        vb = self._prize_board_surface_bg()
        if msg is not None:
            self._update_prize_board_totals_label(None, msg)
            lb = tk.Label(
                ri,
                text=msg,
                font=("Segoe UI", 9),
                bg=vb,
                fg=WHEEL_MUTED,
            )
            lb.pack(anchor=tk.W, padx=4, pady=6)
            return
        if not rows_data:
            self._update_prize_board_totals_label([], None)
            lb = tk.Label(
                ri,
                text="No SKUs with qty left (all are 0).",
                font=("Segoe UI", 9),
                bg=vb,
                fg=WHEEL_MUTED,
            )
            lb.pack(anchor=tk.W, padx=4, pady=6)
            return

        slot, thumb, cols = self._prize_board_slot_thumb_for_inner(ri)
        layout_key = (slot, thumb, cols)
        if self._prize_board_last_layout != layout_key:
            self._inventory_image_cache.clear()
            self._prize_board_last_layout = layout_key

        inset = INV_SLOT_INSET
        inner_side = slot - 2 * inset
        g = INV_GRID_GAP
        sku_font = ("Segoe UI", max(7, min(11, slot // 8)), "bold")
        qty_font = ("Segoe UI", max(9, min(14, slot // 6)), "bold")
        for idx, (sku, q, img) in enumerate(rows_data):
            r, c = divmod(idx, cols)
            outer = tk.Frame(
                ri,
                width=slot,
                height=slot,
                bg=INV_SLOT_EDGE,
                highlightthickness=0,
            )
            outer.grid(row=r, column=c, padx=g, pady=g, sticky=tk.N)
            outer.grid_propagate(False)
            face = tk.Frame(outer, bg=INV_SLOT_FACE)
            face.place(x=inset, y=inset, width=inner_side, height=inner_side)
            refn = _normalize_image_ref(img)
            ph = self._inventory_photo_for_normalized_ref(refn, thumb) if refn else None
            if ph is not None:
                img_lbl = tk.Label(face, image=ph, bg=INV_SLOT_FACE)
                img_lbl.pack(expand=True, pady=(max(2, slot // 14), 2))
            else:
                short = sku if len(sku) <= 12 else sku[:11] + "…"
                tk.Label(
                    face,
                    text=short,
                    font=sku_font,
                    bg=INV_SLOT_FACE,
                    fg=WHEEL_FG,
                    wraplength=max(24, inner_side - 12),
                    justify=tk.CENTER,
                ).pack(expand=True, padx=4, pady=4)
            qty = tk.Label(
                outer,
                text=str(q),
                font=qty_font,
                bg=INV_QTY_BADGE_BG,
                fg=INV_QTY_BADGE_FG,
                padx=max(4, slot // 14),
                pady=1,
            )
            qty.place(relx=1.0, rely=1.0, anchor=tk.SE, x=-3, y=-3)
        self._update_prize_board_totals_label(rows_data, None)

    def _refresh_prizes_label(self) -> None:
        if not hasattr(self, "prizes_label"):
            return
        n = self._total_qty_remaining()
        if n is None:
            text = "Total prizes: —"
        else:
            text = f"Total prizes: {n}"
        self.prizes_label.configure(text=text)
        sc = getattr(self, "_sc_prizes_label", None)
        if sc is not None:
            try:
                if sc.winfo_exists():
                    sc.configure(text=text)
            except tk.TclError:
                pass
        self._update_draw_buttons_for_supply()
        self._reconcile_depleted_wheel_if_applicable()
        self._refresh_remaining_skus_panel()
        self._refresh_overlay_winner_session_list()
        if not self._winner_overlay_poll_started:
            self._winner_overlay_poll_started = True
            self._schedule_winner_overlay_poll()
        self._wheel_publish_html_snapshot()

    def _winner_session_workbook_for_overlay(self) -> Path | None:
        """Workbook shown in the overlay: active session file, else newest under winner_sessions/."""
        p = getattr(self, "_winner_log_path", None)
        if p is not None:
            pp = Path(p)
            if pp.is_file():
                return pp
        base = script_dir() / "winner_sessions"
        return self._find_latest_winner_session_path(base)

    def _schedule_winner_overlay_poll(self) -> None:
        """Poll winner workbook mtime so external edits still refresh the overlay list."""
        aid = self._winner_overlay_poll_after
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
        self._winner_overlay_poll_after = self.after(2500, self._winner_overlay_poll_tick)

    def _winner_overlay_poll_tick(self) -> None:
        self._winner_overlay_poll_after = None
        if hasattr(self, "_overlay_listbox"):
            try:
                self._maybe_refresh_winner_session_overlay_from_mtime()
            except Exception:
                pass
        self._schedule_winner_overlay_poll()

    def _maybe_refresh_winner_session_overlay_from_mtime(self) -> None:
        path = self._winner_session_workbook_for_overlay()
        if path is None or not path.is_file():
            sig: tuple[str, float] = ("", 0.0)
        else:
            try:
                sig = (str(path.resolve()), path.stat().st_mtime)
            except OSError:
                sig = ("", 0.0)
        if sig == self._winner_overlay_last_sig:
            return
        self._winner_overlay_last_sig = sig
        self._refresh_overlay_winner_session_list()

    def _overlay_winner_active_spot(self) -> int | None:
        """Spot # the next SPIN will write to, or the row being filled in fill-skip mode."""
        bf = getattr(self, "_backfill_target_spot", None)
        if bf is not None:
            try:
                return int(bf)
            except (TypeError, ValueError):
                pass
        try:
            return int(self._winner_next_spin)
        except (TypeError, ValueError):
            return None

    def _overlay_spot_target_caption(self) -> str:
        active = self._overlay_winner_active_spot()
        if active is None:
            return ""
        if getattr(self, "_backfill_target_spot", None) is not None:
            return f"▶  Filling Spot {active} in the session sheet (arrow in list below)"
        return f"▶  Next spin writes Spot {active} (arrow in list below)"

    def _apply_overlay_list_active_highlight(self) -> None:
        """Highlight the row for the active spot and scroll it into view."""
        if not hasattr(self, "_overlay_listbox"):
            return
        lb = self._overlay_listbox
        active = self._overlay_winner_active_spot()
        spots = getattr(self, "_overlay_list_line_spots", ())
        active_idx: int | None = None
        if active is not None:
            for i, sp in enumerate(spots):
                if sp == active:
                    active_idx = i
                    break
        try:
            for i in range(lb.size()):
                lb.itemconfig(i, bg=WHEEL_CELL_BG, fg=WHEEL_FG)
            if active_idx is not None:
                lb.itemconfig(active_idx, bg=WHEEL_POINTER, fg="#1a1a2e")
                lb.selection_clear(0, tk.END)
                lb.selection_set(active_idx)
                lb.see(active_idx)
        except tk.TclError:
            pass
        hint = getattr(self, "_overlay_spot_target_label", None)
        if hint is not None:
            try:
                cap = self._overlay_spot_target_caption()
                hint.configure(text=cap if cap else "")
            except tk.TclError:
                pass

    def _refresh_overlay_winner_session_list(self) -> None:
        """Operator-facing rows from the latest / active winner session Excel (main overlay tab).

        Rows are shown in reverse sheet order (typically highest spot / most recent at the top).
        """
        if not hasattr(self, "_overlay_listbox"):
            return
        lb = self._overlay_listbox
        title = getattr(self, "_overlay_list_title", None)
        self._overlay_list_line_spots = []
        active_spot = self._overlay_winner_active_spot()
        try:
            lb.delete(0, tk.END)
            path = self._winner_session_workbook_for_overlay()
            if title is not None:
                if path is not None and path.is_file():
                    title.configure(text=f"Prize Wheel — {path.name}")
                else:
                    title.configure(text="Prize Wheel (no file yet)")
            if load_workbook is None:
                lb.insert(tk.END, "Install openpyxl to view the wheel log.")
                self._set_overlay_winner_list_height()
                self._resize_overlay_to_fit_list()
                self._apply_overlay_list_active_highlight()
                return
            if path is None or not path.is_file():
                lb.insert(tk.END, "(No winner session .xlsx in winner_sessions yet)")
                self._set_overlay_winner_list_height()
                self._resize_overlay_to_fit_list()
                self._apply_overlay_list_active_highlight()
                return
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    if ws is None:
                        lb.insert(tk.END, "(Empty workbook)")
                    else:
                        sheet_rows: list[tuple[int, str, str]] = []
                        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                            a = row[0] if row else None
                            b = row[1] if row and len(row) > 1 else None
                            if a is None and (b is None or str(b).strip() == ""):
                                continue
                            try:
                                if a is not None and str(a).strip() != "":
                                    spot_n = int(a)
                                    spot_s = f"Spot {spot_n}"
                                else:
                                    spot_n = -1
                                    spot_s = str(a).strip() if a is not None else "—"
                            except (TypeError, ValueError):
                                spot_n = -1
                                spot_s = str(a).strip() if a is not None else "—"
                            if b is None or (isinstance(b, str) and not b.strip()) or b == "":
                                prize = "—"
                            else:
                                prize = str(b).strip()
                            sheet_rows.append((spot_n, spot_s, prize))
                        if not sheet_rows:
                            lb.insert(tk.END, "(No winners logged yet in this session)")
                        else:
                            spots_in_sheet = {n for n, _, _ in sheet_rows if n > 0}
                            if (
                                active_spot is not None
                                and active_spot not in spots_in_sheet
                                and getattr(self, "_backfill_target_spot", None) is None
                            ):
                                pending = f"▶  Spot {active_spot}  —  (next spin — not in sheet yet)"
                                lb.insert(tk.END, pending)
                                self._overlay_list_line_spots.append(active_spot)
                            for spot_n, spot_s, prize in reversed(sheet_rows):
                                is_active = active_spot is not None and spot_n == active_spot
                                prefix = "▶  " if is_active else "    "
                                lb.insert(tk.END, f"{prefix}{spot_s}  —  {prize}")
                                self._overlay_list_line_spots.append(
                                    spot_n if spot_n > 0 else None
                                )
                finally:
                    wb.close()
            except Exception as e:
                lb.insert(tk.END, f"Could not read workbook: {e}")
            self._set_overlay_winner_list_height()
            self._resize_overlay_to_fit_list()
            self._apply_overlay_list_active_highlight()
            self._refresh_fill_skipped_hint()
        finally:
            try:
                p2 = self._winner_session_workbook_for_overlay()
                if p2 is not None and p2.is_file():
                    self._winner_overlay_last_sig = (str(p2.resolve()), p2.stat().st_mtime)
                else:
                    self._winner_overlay_last_sig = ("", 0.0)
            except OSError:
                self._winner_overlay_last_sig = ("", 0.0)

    def _resize_overlay_to_fit_list(self) -> None:
        """Resize the overlay so title + capped winner list (see OVERLAY_WINNER_LIST_VISIBLE_ROWS) fit."""
        if not hasattr(self, "_overlay_listbox"):
            return
        try:
            self.update_idletasks()
        except tk.TclError:
            return
        try:
            x, y = self.winfo_rootx(), self.winfo_rooty()
        except tk.TclError:
            return
        w = max(380, self.winfo_reqwidth())
        h = max(200, self.winfo_reqheight())
        try:
            self.geometry(f"{w}x{h}+{x}+{y}")
        except tk.TclError:
            pass

    @staticmethod
    def _wheel_label_text(s: str, max_chars: int = 16) -> str:
        s = (s or "").strip()
        if len(s) <= max_chars:
            return s
        return s[: max_chars - 1] + "…"

    def _wheel_gen_strip(
        self,
        rows: list[tuple[str, int, str]],
        winner: str,
    ) -> tuple[list[str], int, float]:
        self._wheel_dense_pack_active = True
        lap_n = int(type(self).WHEEL_SPIN_LAP_LEN)
        laps = int(type(self).WHEEL_SPIN_LAPS)
        parts: list[str] = []
        for _lap in range(max(1, laps) - 1):
            deco = draw_prize.sample_skus_weighted(rows, 1)
            deco_w = deco[0] if deco else winner
            deco_idx = random.randint(4, max(4, lap_n - 5))
            parts.extend(
                draw_prize.build_wheel_spin_strip(
                    rows, lap_n, winner=deco_w, win_idx=deco_idx
                )
            )
        win_local = random.randint(max(14, lap_n - 10), max(15, lap_n - 3))
        parts.extend(
            draw_prize.build_wheel_spin_strip(
                rows, lap_n, winner=winner, win_idx=win_local
            )
        )
        strip = parts
        physical_win_idx = (max(1, laps) - 1) * lap_n + win_local
        w = max(int(self.wheel_canvas.winfo_width()), 520)
        target = self._wheel_scroll_to_center_index(
            float(physical_win_idx), float(w), dense=True
        )
        self._wheel_note_scroll_viewport(float(w))
        return strip, physical_win_idx, max(0.0, target)

    def _wheel_loading_preview_setup(self, cells: int | None = None) -> None:
        """While waiting on the worker, show random winnable SKUs (potential hits) instead of placeholders."""
        self._wheel_dense_pack_active = True
        self._cancel_wheel_idle_drift()
        if cells is None:
            cells = int(type(self).WHEEL_LOADING_STRIP_LEN)
        self._clear_obs_wheel_result()
        self._wheel_pulse_highlight = False
        fallback = ["···"] * cells
        win_idx = min(cells // 2, max(0, cells - 1))
        cw_px = max(int(self.wheel_canvas.winfo_width()), 520)

        def _center_strip(strip: list[str]) -> None:
            self._wheel_strip = strip
            self._wheel_win_idx = win_idx
            dense = bool(getattr(self, "_wheel_dense_pack_active", False))
            self._wheel_target_scroll = self._wheel_scroll_to_center_index(
                float(win_idx), float(cw_px), dense=dense
            )
            self._wheel_scroll = self._wheel_target_scroll
            self._wheel_note_scroll_viewport(cw_px)

        try:
            p = self._list_path()
        except OSError:
            self._wheel_sku_to_img = {}
            _center_strip(fallback)
            self._wheel_redraw()
            return
        if not p.is_file():
            self._wheel_sku_to_img = {}
            _center_strip(fallback)
            self._wheel_redraw()
            return
        try:
            _, rows = draw_prize.load_rows(p)
        except draw_prize.PrizeDrawError:
            self._wheel_sku_to_img = {}
            _center_strip(fallback)
            self._wheel_redraw()
            return
        self._wheel_sku_to_img = self._sku_to_img_map_from_rows(rows)
        strip = draw_prize.sample_wheel_strip_labels(rows, cells)
        if not strip:
            _center_strip(fallback)
        else:
            _center_strip(strip)
            self._ensure_wheel_photos_for_strip(self._wheel_strip)
        self._wheel_redraw()

    def _wheel_draw_strip_pointer(self, c: tk.Canvas, cx: float, pad_top: float) -> None:
        """Down-pointing indicator above the strip (tip aims into the row)."""
        tip_y = pad_top + 3
        c.create_polygon(
            cx,
            tip_y,
            cx - 11,
            tip_y - 17,
            cx + 11,
            tip_y - 17,
            fill=WHEEL_STRIP_POINTER,
            outline=WHEEL_STRIP_POINTER_EDGE,
            width=1,
        )

    def _wheel_paint_banner_on_canvases(
        self,
        title: str,
        subtitle: str = "",
        *,
        is_error: bool = False,
        title_y_offset: float = 0.0,
        subtitle_y_offset: float = 0.0,
    ) -> None:
        title_fill = WHEEL_ACCENT if is_error else WHEEL_STRIP_WIN_RING
        for c in self._wheel_display_canvases():
            c.delete("all")
            h = int(c.cget("height"))
            w = max(int(c.winfo_width()), 2)
            cx = w / 2.0
            cy = h / 2.0
            c.create_text(
                cx,
                cy + title_y_offset,
                text=title,
                fill=title_fill,
                font=self._title_font,
                width=max(40, w - 40),
            )
            if subtitle:
                c.create_text(
                    cx,
                    cy + subtitle_y_offset,
                    text=subtitle,
                    fill=WHEEL_STRIP_FG,
                    font=("Segoe UI", 10),
                    width=max(48, w - 48),
                )
            self._wheel_draw_strip_pointer(c, cx, float(WHEEL_STRIP_PAD_TOP))

    def _wheel_redraw_to_canvas(self, c: tk.Canvas, *, _idle_drift_frame: bool = False) -> None:
        c.delete("all")
        h = int(c.cget("height"))
        w = max(int(c.winfo_width()), 2)
        cx = w / 2.0
        scroll_vis = self._wheel_scroll_for_strip_render_on_canvas(c)
        scroll_model = self._wheel_scroll_for_canvas(c, float(self._wheel_scroll))
        target_canvas = self._wheel_scroll_for_canvas(c, float(self._wheel_target_scroll))
        cw = float(self.WHEEL_CELL)
        dense = bool(getattr(self, "_wheel_dense_pack_active", False))
        pl, pr = self._wheel_cell_xmargins(dense=dense)
        pad_top = float(WHEEL_STRIP_PAD_TOP)
        pad_bot = float(WHEEL_STRIP_PAD_BOTTOM)
        cell_h = max(32.0, float(h) - pad_top - pad_bot)
        landed = abs(scroll_model - target_canvas) < 1.5
        gap = float(type(self).WHEEL_SPIN_CELL_GAP if dense else WHEEL_STRIP_CELL_GAP)

        y_track = min(float(h) - 3.0, pad_top + cell_h + 3.0)
        c.create_line(0, y_track, w, y_track, fill=WHEEL_STRIP_TRACK, width=1)

        n_idle_phys = int(type(self).WHEEL_IDLE_STRIP_LEN) * int(type(self).WHEEL_IDLE_CAROUSEL_COPIES)
        idle_carousel = len(self._wheel_strip) == n_idle_phys
        edge_pad = cw if dense else 0.0
        for i, raw in enumerate(self._wheel_strip):
            x0 = i * cw - scroll_vis + pl
            x1 = i * cw - scroll_vis + cw - pr
            if x1 < -edge_pad or x0 > w + edge_pad:
                continue
            rx0, rx1 = x0 + gap, x1 - gap
            if rx1 <= rx0 + 1.0:
                continue
            cell_center_x = i * cw + (pl + cw - pr) / 2.0 - scroll_vis
            if idle_carousel:
                win_cell = False
            elif len(self._wheel_strip) == self.WHEEL_IDLE_STRIP_LEN:
                win_cell = landed and abs(cell_center_x - cx) < cw * 0.38
            else:
                win_cell = landed and i == self._wheel_win_idx and abs(cell_center_x - cx) < cw * 0.38
            outline_w = 1
            if win_cell:
                outline_w = 3 if self._wheel_pulse_highlight else 2
            fill_c = WHEEL_STRIP_WIN_BG if win_cell else WHEEL_STRIP_SLOT_BG
            outline_c = WHEEL_STRIP_WIN_RING if win_cell else WHEEL_STRIP_SLOT_BD
            fg_c = WHEEL_STRIP_WIN_FG if win_cell else WHEEL_STRIP_FG
            c.create_rectangle(
                rx0,
                pad_top + 2,
                rx1,
                pad_top + cell_h - 2,
                fill=fill_c,
                outline=outline_c,
                width=outline_w,
            )
            img_ref = _normalize_image_ref(self._wheel_sku_to_img.get(raw, ""))
            ph = self._wheel_image_cache.get(img_ref) if img_ref else None
            cx_cell = (rx0 + rx1) / 2.0
            cy_cell = pad_top + 2.0 + (cell_h - 4.0) / 2.0
            slot_font = self._wheel_font_bold if win_cell else self._wheel_font
            if ph is not None:
                c.create_image(cx_cell, cy_cell, image=ph)
            else:
                c.create_text(
                    cx_cell,
                    cy_cell,
                    text=self._wheel_label_text(raw),
                    fill=fg_c,
                    font=slot_font,
                    width=int(max(12.0, rx1 - rx0 - 12.0)),
                )

        self._wheel_draw_strip_pointer(c, cx, pad_top)

    def _wheel_redraw(self, _idle_drift_frame: bool = False) -> None:
        if len(self._wheel_strip) > 0:
            self._wheel_banner_title = None
            self._wheel_banner_subtitle = None
            self._wheel_banner_is_error = False
        for c in self._wheel_display_canvases():
            self._wheel_redraw_to_canvas(c, _idle_drift_frame=_idle_drift_frame)
        self._wheel_publish_html_snapshot()
        if (
            not _idle_drift_frame
            and self._wheel_idle_drift_active()
            and self._wheel_idle_drift_after is None
        ):
            self._wheel_idle_drift_after = self.after(
                int(type(self).WHEEL_IDLE_DRIFT_MS),
                self._wheel_idle_drift_tick,
            )

    def _reset_wheel_idle(self) -> None:
        self._cancel_pulse()
        self._cancel_wheel_idle_drift()
        self._clear_obs_wheel_result()
        self._wheel_pulse_highlight = False
        self._wheel_dense_pack_active = False
        units = self._winnable_units_total()
        if units is not None and units == 0:
            self._update_draw_buttons_for_supply()
            self._reconcile_depleted_wheel_if_applicable()
            return
        self._customer_empty_wheel = False
        self._paint_normal_idle_wheel_only()
        self._update_draw_buttons_for_supply()

    def _hide_super_reroll_keep_buttons(self) -> None:
        """Super-only actions stay out of the draw bar until a Super spin is waiting for Reroll/Keep."""
        if not hasattr(self, "reroll_btn"):
            return
        for w in (self.reroll_btn, self.keep_btn):
            try:
                w.pack_forget()
            except tk.TclError:
                pass

    def _show_super_reroll_keep_buttons(self) -> None:
        if not hasattr(self, "reroll_btn"):
            return
        self.reroll_btn.pack(side=tk.LEFT, padx=(0, 10), after=self.super_btn)
        self.keep_btn.pack(side=tk.LEFT, padx=(0, 10), after=self.reroll_btn)

    def _style_main_controls_idle(self) -> None:
        if not hasattr(self, "spin_btn"):
            return
        self._hide_super_reroll_keep_buttons()
        self.spin_btn.config(state=tk.NORMAL, bg=BTN_SPIN)
        self.super_btn.config(state=tk.NORMAL, bg=BTN_SUPER)
        self.skip_spot_btn.config(state=tk.NORMAL, bg=BTN_SKIP)
        self._sync_undo_spin_button()

    def _style_super_controls_active(self) -> None:
        if not hasattr(self, "spin_btn"):
            return
        self.spin_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self.super_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        self._show_super_reroll_keep_buttons()
        if self._super_reroll_used:
            self.reroll_btn.config(state=tk.DISABLED, bg=BTN_DISABLED)
        else:
            self.reroll_btn.config(state=tk.NORMAL, bg=BTN_REROLL)
        self.keep_btn.config(state=tk.NORMAL, bg=BTN_KEEP)
        self.skip_spot_btn.config(state=tk.NORMAL, bg=BTN_SKIP)
        self._sync_undo_spin_button()

    def _wheel_show_error(self, message: str) -> None:
        self._cancel_wheel_idle_drift()
        self._clear_obs_wheel_result()
        self._wheel_sku_to_img = {}
        self._wheel_strip = []
        self._wheel_banner_title = str(message)
        self._wheel_banner_subtitle = ""
        self._wheel_banner_is_error = True
        self._wheel_paint_banner_on_canvases(message, is_error=True, title_y_offset=6.0)
        self._set_wheel_status(message, WHEEL_ACCENT)
        self._style_main_controls_idle()
        self._wheel_publish_html_snapshot()

    def _cancel_pulse(self) -> None:
        if self._pulse_after is not None:
            try:
                self.after_cancel(self._pulse_after)
            except tk.TclError:
                pass
            self._pulse_after = None

    def _pulse_wheel_win(self, flashes: int = 10) -> None:
        self._cancel_pulse()
        state = [0]

        def step() -> None:
            state[0] += 1
            if state[0] > flashes:
                self._wheel_pulse_highlight = False
                self._wheel_redraw()
                self._pulse_after = None
                self._update_draw_buttons_for_supply()
                self._reconcile_depleted_wheel_if_applicable()
                return
            self._wheel_pulse_highlight = state[0] % 2 == 1
            self._wheel_redraw()
            self._pulse_after = self.after(110, step)

        step()

    def _run_wheel_spin(
        self,
        rows: list[tuple[str, int, str]],
        winner: str,
        subtitle: str,
        done: object,
        sku_to_img: dict[str, str] | None = None,
    ) -> None:
        self._clear_obs_wheel_result()
        self._wheel_banner_title = None
        self._wheel_banner_subtitle = None
        self._wheel_banner_is_error = False
        self._clear_animation()
        self._cancel_pulse()
        self._cancel_wheel_idle_drift()
        self._wheel_strip, self._wheel_win_idx, self._wheel_target_scroll = self._wheel_gen_strip(
            rows, winner
        )
        self._wheel_sku_to_img = dict(sku_to_img) if sku_to_img else {}
        self._ensure_wheel_photos_for_strip(self._wheel_strip)
        self._wheel_scroll = 0.0
        self._wheel_note_scroll_viewport(self._wheel_primary_viewport_w())
        self._wheel_pulse_highlight = False
        self._set_wheel_status("Wheel spinning…", WHEEL_TITLE)
        self._wheel_redraw()

        start = time.monotonic()
        dur = self._wheel_spin_duration_sec()
        tick_ms = int(type(self).WHEEL_SPIN_TICK_MS)
        target = self._wheel_target_scroll

        def tick() -> None:
            elapsed = time.monotonic() - start
            t = min(1.0, elapsed / dur)
            ease = type(self)._wheel_spin_ease(t)
            self._wheel_scroll = target * ease
            self._wheel_redraw()
            if t >= 1.0:
                self._anim_after = None
                self._set_wheel_status(subtitle, WHEEL_FG)
                done()
            else:
                self._anim_after = self.after(tick_ms, tick)

        self._anim_after = self.after(tick_ms, tick)

    def _list_path(self) -> Path:
        t = self.path_var.get().strip()
        if not t:
            raise OSError("Prize list path is not set")
        return Path(t).expanduser().resolve()

    def _maybe_refresh_idle_wheel(self) -> None:
        """Re-paint the idle preview strip when the list path or stock changes (not during spin/super)."""
        if (
            hasattr(self, "wheel_canvas")
            and not self._busy
            and self._pending_super is None
            and self._pulse_after is None
            and self._anim_after is None
        ):
            self._reset_wheel_idle()

    def _on_path_write(self, *_args: object) -> None:
        try:
            key = str(self._list_path())
        except OSError:
            self._clear_undo_spin()
            self._refresh_prizes_label()
            self._maybe_refresh_idle_wheel()
            self._refresh_setup_status()
            return
        if self._session_target and key != self._session_target:
            self._hide_super_panel()
            self._invalidate_session()
        self._clear_undo_spin()
        self._refresh_prizes_label()
        self._maybe_refresh_idle_wheel()
        self._refresh_setup_status()
        if self._setup_expanded.get():
            self.after_idle(self._overlay_resize_for_setup)
        else:
            self.after_idle(self._setup_sync_scroll_region)
        self._schedule_save_app_state()

    def _log(self, text: str) -> None:
        self.out.insert(tk.END, text)
        self.out.see(tk.END)

    @staticmethod
    def _find_latest_winner_session_path(base: Path) -> Path | None:
        if not base.is_dir():
            return None
        paths = [p for p in base.glob("*.xlsx") if p.is_file()]
        if not paths:
            return None
        return max(paths, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _winner_session_stem_matches_wheel(stem: str, slug: str) -> bool:
        """True if an existing winner_sessions .xlsx belongs to this wheel id."""
        s = (stem or "").casefold()
        slug_cf = (slug or "").casefold()
        if not s or not slug_cf:
            return False
        return (
            s.startswith(f"{slug_cf}_winners")
            or s.startswith(f"winners_{slug_cf}_")
            or s == slug_cf
        )

    def _find_existing_winner_session_for_wheel(
        self, wid: str, base: Path | None = None
    ) -> Path | None:
        """Newest winner_sessions workbook for this wheel, if any."""
        slug = self._wheel_name_slug_for_filename(wid)
        if not slug:
            return None
        root = (base if base is not None else script_dir() / "winner_sessions").resolve()
        if not root.is_dir():
            return None
        candidates: list[Path] = []
        for p in root.glob("*.xlsx"):
            if not p.is_file():
                continue
            if not self._winner_session_stem_matches_wheel(p.stem, slug):
                continue
            try:
                self._next_winner_spot_from_workbook(p)
            except Exception:
                continue
            candidates.append(p)
        if not candidates:
            return None
        return max(candidates, key=lambda p: p.stat().st_mtime)

    @classmethod
    def _winner_session_filename_stamp(cls, dt: datetime | None = None) -> str:
        """Month_day and hour-minute AM/PM in Texas (Central) time, e.g. May_16_3-45PM."""
        when = dt or datetime.now(cls.TEXAS_TZ)
        if when.tzinfo is None:
            when = when.replace(tzinfo=cls.TEXAS_TZ)
        else:
            when = when.astimezone(cls.TEXAS_TZ)
        month_day = f"{when.strftime('%B')}_{when.day}"
        hour = int(when.strftime("%I"))
        minute = when.strftime("%M")
        ampm = when.strftime("%p")
        return f"{month_day}_{hour}-{minute}{ampm}"

    @classmethod
    def _winner_session_workbook_basename(cls, wheel_id_slug: str) -> str:
        """e.g. Wheel58912_winners_May_16_3-45PM"""
        slug = (wheel_id_slug or "").strip() or "wheel"
        return f"{slug}_winners_{cls._winner_session_filename_stamp()}"

    @staticmethod
    def _wheel_name_slug_for_filename(raw: str, max_len: int = 40) -> str:
        """ASCII slug for <wheel>_winners_<date_time>.xlsx; empty string if nothing usable remains."""
        s = (raw or "").strip()
        if not s:
            return ""
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"[^0-9A-Za-z_-]+", "", s)
        s = s.strip("_-")
        if not s:
            return ""
        if len(s) > max_len:
            s = s[:max_len].rstrip("_-.")
        return s

    @staticmethod
    def _parse_spot_number_from_label(label: str) -> int | None:
        m = re.search(r"Spot\s+(\d+)", (label or "").strip(), re.I)
        if not m:
            return None
        try:
            return int(m.group(1))
        except ValueError:
            return None

    def _winner_assignment_display_for_spot(self, spot_num: int) -> str:
        """Prize (SKU) text for this spot # from the session winner workbook, for OBS overlay."""
        path = getattr(self, "_winner_log_path", None)
        if load_workbook is None:
            return "Install openpyxl for winner log"
        if path is None:
            return "No winner session file"
        p = Path(path)
        if not p.is_file():
            return "No winner session file"
        try:
            mtime = p.stat().st_mtime
        except OSError:
            return "Winner log unreadable"
        cache = getattr(self, "_winner_spot_lookup_cache", None)
        if cache is not None:
            c_path, c_mtime, c_spot, c_text = cache
            if c_path == str(p.resolve()) and c_mtime == mtime and c_spot == spot_num:
                return c_text
        found = False
        sku_cell: str | None = None
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
            try:
                ws = wb.active
                if ws is None:
                    txt = "Winner sheet empty"
                    self._winner_spot_lookup_cache = (str(p.resolve()), mtime, spot_num, txt)
                    return txt
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    a = row[0] if row else None
                    if a is None or a == "":
                        continue
                    try:
                        n = int(a)
                    except (TypeError, ValueError):
                        continue
                    if n != spot_num:
                        continue
                    found = True
                    b = row[1] if row and len(row) > 1 else None
                    if b is None or (isinstance(b, str) and not b.strip()) or b == "":
                        sku_cell = ""
                    else:
                        sku_cell = str(b).strip()
                    break
            finally:
                wb.close()
        except Exception:
            return "Winner log unreadable"
        if not found:
            txt = ""
        elif not sku_cell:
            txt = "No prize (skipped / empty)"
        else:
            txt = sku_cell
        self._winner_spot_lookup_cache = (str(p.resolve()), mtime, spot_num, txt)
        return txt

    @staticmethod
    def _next_winner_spot_from_workbook(path: Path) -> int:
        """Next spot # after the highest integer in column A (rows below header)."""
        if load_workbook is None:
            return 1
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return 1
            m = 0
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                val = row[0] if row else None
                if val is None or val == "":
                    continue
                try:
                    n = int(val)
                except (TypeError, ValueError):
                    continue
                if n > m:
                    m = n
            return m + 1 if m else 1
        finally:
            wb.close()

    def list_skipped_spots_eligible_for_fill(self) -> list[int]:
        """Spot numbers in the active winner session that have a row with an empty prize (skipped / pending).

        Use with ``prepare_spin_for_skipped_spot`` so the next normal SPIN writes the prize into that row
        instead of appending a new spot. After a successful fill, the app returns to the usual next spot.
        """
        path = getattr(self, "_winner_log_path", None)
        if path is None or load_workbook is None:
            return []
        p = Path(path)
        if not p.is_file():
            return []
        out: list[int] = []
        seen: set[int] = set()
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
            try:
                ws = wb.active
                if ws is None:
                    return []
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    a = row[0] if row else None
                    b = row[1] if row and len(row) > 1 else None
                    if a is None or str(a).strip() == "":
                        continue
                    try:
                        n = int(a)
                    except (TypeError, ValueError):
                        continue
                    if n in seen:
                        continue
                    b_s = "" if b is None else str(b).strip()
                    if b_s != "":
                        continue
                    seen.add(n)
                    out.append(n)
            finally:
                wb.close()
        except Exception:
            return []
        out.sort()
        return out

    def _winner_row_has_empty_prize(self, workbook_path: Path, spot_number: int) -> bool:
        if load_workbook is None:
            return False
        try:
            wb = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                ws = wb.active
                if ws is None:
                    return False
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    a = row[0] if row else None
                    b = row[1] if row and len(row) > 1 else None
                    if a is None or str(a).strip() == "":
                        continue
                    try:
                        n = int(a)
                    except (TypeError, ValueError):
                        continue
                    if n != spot_number:
                        continue
                    b_s = "" if b is None else str(b).strip()
                    return b_s == ""
            finally:
                wb.close()
        except Exception:
            return False
        return False

    def prepare_spin_for_skipped_spot(self, spot_number: int) -> str | None:
        """Arm the next **normal SPIN** so its prize is written into an existing skipped row for ``spot_number``.

        Validates against the active Energy Break winner session workbook (empty prize cell for that spot).
        On success returns ``None``; on failure returns a short error message. Does not run the wheel — press SPIN.

        Call ``cancel_fill_skipped_spot_mode`` to return to the usual next spot without spinning.
        """
        if self._busy or self._pending_super is not None:
            return "Finish the current spin (including Super / Reroll / Keep) first."
        if load_workbook is None:
            return "Install openpyxl to use the winner session file."
        path = getattr(self, "_winner_log_path", None)
        if path is None:
            return "No winner session file is configured."
        p = Path(path)
        if not p.is_file():
            return "Winner session file is missing on disk."
        if not self._winner_row_has_empty_prize(p, int(spot_number)):
            return (
                f"There is no row for Spot #{spot_number} with an empty prize in this session. "
                "Check the winner list or pick another spot."
            )
        self._backfill_target_spot = int(spot_number)
        self._update_spin_counter_label()
        self._refresh_overlay_winner_session_list()
        self._wheel_publish_html_snapshot()
        nxt = self._winner_next_spin
        self._log(
            f"Fill-skip mode: next SPIN writes the prize into Spot #{spot_number}'s empty row. "
            f"After that, the wheel goes back to the normal queue (next new row would be #{nxt}).\n"
        )
        return None

    def cancel_fill_skipped_spot_mode(self, *, silent: bool = False) -> None:
        """Leave fill-skip mode without spinning; the counter returns to the usual next spot."""
        if getattr(self, "_backfill_target_spot", None) is None:
            return
        self._backfill_target_spot = None
        self._update_spin_counter_label()
        self._refresh_overlay_winner_session_list()
        self._wheel_publish_html_snapshot()
        self._set_fill_skipped_status("")
        if not silent:
            self._log("Fill-skip mode cancelled — using normal next spot again.\n")

    def _refresh_fill_skipped_hint(self) -> None:
        lab = getattr(self, "_fill_skipped_hint_label", None)
        if lab is None:
            return
        try:
            spots = self.list_skipped_spots_eligible_for_fill()
        except Exception:
            spots = []
        if spots:
            shown = ", ".join(f"#{s}" for s in spots[:16])
            extra = f" (+{len(spots) - 16} more)" if len(spots) > 16 else ""
            lab.configure(text=f"Empty prize in sheet: {shown}{extra}")
        else:
            lab.configure(text="No skipped spots with an empty prize in the active session file.")

    def _set_fill_skipped_status(self, text: str, *, ok: bool = True) -> None:
        lab = getattr(self, "_fill_skipped_status_label", None)
        if lab is None:
            return
        try:
            lab.configure(text=text, fg=WHEEL_POINTER if ok else WHEEL_ACCENT)
        except tk.TclError:
            pass

    def _open_fill_skipped_spot_dialog(self) -> None:
        """Pick a skipped spot from the winner session, then use normal SPIN to write the prize into that row."""
        if self._busy or self._pending_super is not None:
            messagebox.showinfo("Fill skipped spot", "Finish the current spin first.")
            return
        block = self._winner_session_spin_block_reason()
        if block:
            messagebox.showinfo("Fill skipped spot", block)
            return
        spots = self.list_skipped_spots_eligible_for_fill()
        if not spots:
            messagebox.showinfo(
                "Fill skipped spot",
                "No rows with an empty prize were found in the active winner session file.\n\n"
                "Skipped spots must appear in the sheet with column B empty.",
            )
            return
        top = tk.Toplevel(self)
        top.title("Fill a skipped spot")
        top.configure(bg=PRIZE_BOARD_CONTENT_BG)
        top.transient(self)
        top.grab_set()
        tk.Label(
            top,
            text="Choose a spot that was skipped (empty prize). Then press SPIN on the draw controls.",
            bg=PRIZE_BOARD_CONTENT_BG,
            fg=WHEEL_FG,
            font=("Segoe UI", 10),
            wraplength=400,
            justify=tk.LEFT,
        ).pack(fill=tk.X, padx=14, pady=(14, 8))
        lb = tk.Listbox(
            top,
            height=min(12, max(4, len(spots))),
            font=("Segoe UI", 11),
            bg=INV_SLOT_FACE,
            fg=WHEEL_FG,
            selectbackground=WHEEL_POINTER,
            selectforeground="#1a1a2e",
            highlightthickness=0,
        )
        for s in spots:
            lb.insert(tk.END, f"Spot {s}  —  (empty)")
        lb.pack(fill=tk.BOTH, expand=True, padx=14, pady=4)
        if spots:
            lb.selection_set(0)
            lb.activate(0)
        btn_row = tk.Frame(top, bg=PRIZE_BOARD_CONTENT_BG)
        btn_row.pack(fill=tk.X, padx=14, pady=(4, 14))

        def arm() -> None:
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo("Fill skipped spot", "Select a spot in the list first.", parent=top)
                return
            spot_n = spots[int(sel[0])]
            err = self.prepare_spin_for_skipped_spot(spot_n)
            if err:
                messagebox.showerror("Fill skipped spot", err, parent=top)
                return
            _close_fill_skipped_dialog(top)
            self._set_fill_skipped_status(
                f"Spot #{spot_n} armed — press SPIN once. Then the wheel returns to spot #{self._winner_next_spin}.",
                ok=True,
            )

        lb.bind("<Double-Button-1>", lambda _e: arm())
        tk.Button(
            btn_row,
            text="  Use this spot for next SPIN  ",
            command=arm,
            font=("Segoe UI", 10, "bold"),
            bg=BTN_SPIN,
            fg="white",
            relief=tk.FLAT,
            padx=12,
            pady=8,
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(
            btn_row,
            text="  Close  ",
            command=lambda: _close_fill_skipped_dialog(top),
            font=("Segoe UI", 10),
            bg=WHEEL_CELL_BG,
            fg=WHEEL_FG,
            relief=tk.FLAT,
            padx=12,
            pady=8,
            cursor="hand2",
        ).pack(side=tk.LEFT)

    def _update_winner_row_prize_for_spot(self, spot: int, sku: str) -> bool:
        """Set column B for the first row whose column A equals ``spot`` and B is empty."""
        path = self._winner_log_path
        if path is None or load_workbook is None:
            return False
        p = Path(path)
        wb = None
        try:
            wb = load_workbook(p)
            ws = wb.active
            if ws is None:
                return False
            max_r = ws.max_row or 1
            sku_clean = (sku or "").strip()
            if not sku_clean:
                return False
            for r in range(2, max_r + 1):
                a = ws.cell(row=r, column=1).value
                try:
                    n = int(a) if a is not None and str(a).strip() != "" else None
                except (TypeError, ValueError):
                    n = None
                if n != spot:
                    continue
                b = ws.cell(row=r, column=2).value
                b_s = "" if b is None else str(b).strip()
                if b_s != "":
                    continue
                ws.cell(row=r, column=2, value=sku_clean)
                wb.save(p)
                return True
            return False
        except OSError:
            return False
        except Exception:
            return False
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def _winner_clear_prize_cell_for_spot(self, path: Path, spot_number: int) -> bool:
        """Clear column B (empty prize) for the last data row whose column A equals ``spot_number``."""
        if load_workbook is None:
            return False
        wb = None
        try:
            wb = load_workbook(path)
            ws = wb.active
            if ws is None:
                return False
            target: int | None = None
            max_r = ws.max_row or 1
            for r in range(2, max_r + 1):
                a = ws.cell(row=r, column=1).value
                try:
                    n = int(a) if a is not None and str(a).strip() != "" else None
                except (TypeError, ValueError):
                    n = None
                if n == spot_number:
                    target = r
            if target is None:
                return False
            ws.cell(row=target, column=2, value="")
            wb.save(path)
            return True
        except Exception:
            return False
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def _write_new_winner_workbook_file(self, path: Path) -> bool:
        if Workbook is None:
            return False
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise RuntimeError("No active worksheet")
            ws.title = "Winners"
            ws.append(["Spot #", "Prize (SKU)"])
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 44
            wb.save(path)
            return True
        except OSError:
            return False

    def _create_new_winner_workbook(self, base: Path, wheel_name_slug: str = "") -> Path | None:
        """Create <wheel>_winners_<Month>_<day>_<h-MMAMPM>.xlsx (Texas Central time)."""
        if Workbook is None:
            return None
        slug = self._wheel_name_slug_for_filename(wheel_name_slug) or "wheel"
        basename = self._winner_session_workbook_basename(slug)
        path = base / f"{basename}.xlsx"
        n = 2
        while path.is_file():
            path = base / f"{basename}_{n}.xlsx"
            n += 1
        return path if self._write_new_winner_workbook_file(path) else None

    def _winner_log_path_display(self) -> str:
        wp = getattr(self, "_winner_log_path", None)
        if wp is not None:
            try:
                p = Path(wp)
                if p.is_file():
                    return f"Active spreadsheet: {p.name}"
            except OSError:
                pass
        return "Active spreadsheet: none — choose a file before spinning"

    def _refresh_winner_log_path_label(self) -> None:
        lab = getattr(self, "_winner_log_path_label", None)
        if lab is None:
            return
        try:
            if lab.winfo_exists():
                lab.configure(text=self._winner_log_path_display())
        except tk.TclError:
            pass

    def _attach_winner_session_path(self, path: Path) -> Path:
        """Use a winner session workbook from persisted settings (no filename mapping)."""
        p = path.expanduser().resolve()
        self._winner_log_path = p
        try:
            self._winner_next_spin = self._next_winner_spot_from_workbook(p)
        except Exception:
            self._winner_next_spin = 1
        self._clear_undo_spin()
        self._update_spin_counter_label()
        self._refresh_overlay_winner_session_list()
        self._refresh_winner_log_path_label()
        self._schedule_save_app_state()
        return p

    def _create_and_attach_winner_session_for_preset(self, wid: str) -> Path | None:
        """Reuse an existing winner_sessions file for this wheel, or create a new one."""
        if Workbook is None:
            self._log(
                f"Wheel preset {wid}: install openpyxl to create a winner spreadsheet.\n"
            )
            return None
        base = script_dir() / "winner_sessions"
        try:
            base.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            self._log(f"Wheel preset {wid}: could not create winner_sessions folder: {e}\n")
            return None
        existing = self._find_existing_winner_session_for_wheel(wid, base)
        if existing is not None:
            self._attach_winner_session_path(existing.resolve())
            self._log(
                f"Winner spreadsheet for preset {wid}: reusing {existing.name} "
                f"(next spot #{self._winner_next_spin}).\n"
            )
            return existing
        slug = self._wheel_name_slug_for_filename(wid)
        path = self._create_new_winner_workbook(base, slug)
        if path is None:
            self._log(f"Wheel preset {wid}: could not create a new winner spreadsheet.\n")
            return None
        self._attach_winner_session_path(path.resolve())
        self._log(f"New winner spreadsheet for preset {wid}: {path.name}\n")
        return path

    def _choose_winner_session_spreadsheet(self) -> None:
        """Pick an existing .xlsx file to receive spin / skip rows."""
        if Workbook is None or load_workbook is None:
            messagebox.showinfo(
                "Winner spreadsheet",
                "Install openpyxl to use winner session Excel files.",
            )
            return
        if getattr(self, "_backfill_target_spot", None) is not None:
            messagebox.showinfo(
                "Winner spreadsheet",
                "Cancel fill-skip mode first (Cancel fill mode in Spin & controls).",
            )
            return
        if self._busy or self._pending_super is not None:
            messagebox.showinfo(
                "Winner spreadsheet",
                "Finish or cancel the current spin (including Super / Reroll / Keep) first.",
            )
            return
        base = script_dir() / "winner_sessions"
        initial = base if base.is_dir() else script_dir()
        wp = getattr(self, "_winner_log_path", None)
        if wp is not None:
            try:
                parent = Path(wp).parent
                if parent.is_dir():
                    initial = parent
            except OSError:
                pass
        picked = filedialog.askopenfilename(
            title="Choose winner spreadsheet to populate",
            initialdir=str(initial),
            filetypes=[
                ("Excel workbook", "*.xlsx"),
                ("All files", "*.*"),
            ],
        )
        if not picked:
            return
        path = Path(picked).expanduser().resolve()
        if not path.is_file():
            messagebox.showerror("Winner spreadsheet", "That file does not exist.")
            return
        try:
            self._next_winner_spot_from_workbook(path)
        except Exception as e:
            messagebox.showerror(
                "Winner spreadsheet",
                f"Could not read that workbook:\n{e}\n\n"
                "Use a file with Spot # and Prize (SKU) columns (row 1 = headers).",
            )
            return
        self._attach_winner_session_path(path)
        self._log(
            f"Winner spreadsheet set to {path.name} (next spot #{self._winner_next_spin}).\n"
        )
        self._update_draw_buttons_for_supply()

    def _collect_app_state(self) -> dict[str, str]:
        out: dict[str, str] = {}
        preset = (self.wheel_preset_var.get() or getattr(self, "_active_wheel_preset_id", "") or "").strip()
        if preset:
            out["wheel_preset_id"] = preset
        list_raw = self.path_var.get().strip()
        if list_raw:
            out["list_path"] = list_raw
        images_raw = self.images_path_var.get().strip()
        if images_raw:
            out["images_path"] = images_raw
        wp = getattr(self, "_winner_log_path", None)
        if wp is not None:
            out["winner_log_path"] = str(Path(wp).resolve())
        return out

    def _schedule_save_app_state(self) -> None:
        aid = getattr(self, "_app_state_save_after", None)
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
        self._app_state_save_after = self.after(350, self._flush_save_app_state)

    def _flush_save_app_state(self) -> None:
        self._app_state_save_after = None
        self._save_app_state()

    def _save_app_state(self) -> None:
        data = self._collect_app_state()
        if not data:
            return
        if app_local_state.save_state(script_dir(), data):
            return
        self._log("Warning: could not save session state to energy_break_state.json\n")

    def _load_app_state(self) -> bool:
        raw = app_local_state.load_state(script_dir())
        if not raw:
            return False
        restored = False
        list_raw = str(raw.get("list_path") or "").strip()
        if list_raw:
            self.path_var.set(list_raw)
            restored = True
        images_raw = str(raw.get("images_path") or "").strip()
        if images_raw:
            self.images_path_var.set(images_raw)
            restored = True
        preset = str(raw.get("wheel_preset_id") or "").strip()
        if preset:
            self.wheel_preset_var.set(preset)
            self._active_wheel_preset_id = draw_prize.normalize_wheel_id(preset)
            restored = True
        winner_raw = str(raw.get("winner_log_path") or "").strip()
        preset_norm = draw_prize.normalize_wheel_id(preset) if preset else ""
        if winner_raw:
            wp = Path(winner_raw)
            if wp.is_file():
                self._attach_winner_session_path(wp)
                restored = True
            else:
                self._winner_log_path = None
                if preset_norm:
                    self._create_and_attach_winner_session_for_preset(preset_norm)
                    restored = True
                else:
                    self._log(
                        f"Saved winner log not found ({wp.name}) — apply a preset or choose a spreadsheet.\n"
                    )
        elif preset_norm:
            self._create_and_attach_winner_session_for_preset(preset_norm)
            restored = True
        if restored:
            self._log("Restored last session from energy_break_state.json\n")
            self._invalidate_session()
            self._on_images_path_write()
            self._refresh_setup_status()
            self.after_idle(self._update_draw_buttons_for_supply)
        return restored

    def _init_winner_session_log(self) -> None:
        """No winner log until restored from energy_break_state.json or a new preset is applied."""
        self._winner_next_spin = 1
        self._winner_log_path = None
        try:
            (script_dir() / "winner_sessions").mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
        self.after_idle(self._update_draw_buttons_for_supply)

    def _append_winner_sheet_row(self, sku: str | None) -> bool:
        """Append a new row, or (fill-skip mode) update column B on an existing skipped row."""
        path = self._winner_log_path
        if path is None or load_workbook is None:
            try:
                self._refresh_overlay_winner_session_list()
            except Exception:
                pass
            return False
        bf = getattr(self, "_backfill_target_spot", None)
        if bf is not None and sku is not None and str(sku).strip() != "":
            ok = self._update_winner_row_prize_for_spot(int(bf), str(sku).strip())
            if ok and hasattr(self, "_winner_spot_lookup_cache"):
                self._winner_spot_lookup_cache = None
            try:
                self._refresh_overlay_winner_session_list()
            except Exception:
                pass
            return ok
        sku_cell = "" if sku is None else sku
        ok = False
        wb = None
        try:
            wb = load_workbook(path)
            ws = wb.active
            if ws is None:
                raise RuntimeError("No active worksheet")
            ws.append([self._winner_next_spin, sku_cell])
            wb.save(path)
            self._winner_next_spin += 1
            self._update_spin_counter_label()
            ok = True
        except OSError as e:
            self._log(f"Warning: could not append to winner list: {e}\n")
        except Exception as e:
            self._log(f"Warning: could not append to winner list ({type(e).__name__}): {e}\n")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            try:
                self._refresh_overlay_winner_session_list()
            except Exception:
                pass
        return ok

    def _record_spin_winner(self, sku: str) -> bool:
        return self._append_winner_sheet_row(sku)

    def _update_spin_counter_label(self) -> None:
        if not hasattr(self, "spin_counter_label"):
            return
        bf = getattr(self, "_backfill_target_spot", None)
        if bf is not None:
            text = f"Spot {bf} — fill skipped row"
        else:
            text = f"Spot {self._winner_next_spin}"
        self.spin_counter_label.configure(text=text)
        sc = getattr(self, "_sc_spin_counter_label", None)
        if sc is not None:
            try:
                if sc.winfo_exists():
                    sc.configure(text=text)
            except tk.TclError:
                pass
        if hasattr(self, "_overlay_listbox"):
            try:
                self._apply_overlay_list_active_highlight()
            except tk.TclError:
                pass

    def _on_skip_spot(self) -> None:
        if self._busy:
            return
        block = self._winner_session_spin_block_reason()
        if block:
            messagebox.showinfo("Cannot skip spot", block)
            return
        if getattr(self, "_backfill_target_spot", None) is not None:
            messagebox.showinfo(
                "Skip spot",
                "Cancel \"Fill a skipped spot\" first (Cancel fill mode in Spin & controls).",
            )
            return
        n = self._winner_next_spin
        if not messagebox.askyesno(
            "Skip spot",
            f"A row will be added for spot #{n} with an empty prize (no SKU). "
            f"The next row will use #{n + 1}. Use this when a buyer's payment did not clear.",
        ):
            return
        if self._append_winner_sheet_row(None):
            self._log(f"Skipped spot #{n}: row added with empty SKU. Next winner list #: {self._winner_next_spin}.\n")
            self._push_undo_skip_log(n)
        else:
            self._winner_next_spin += 1
            self._update_spin_counter_label()
            self._log(
                f"Skipped spot #{n} (winner file unavailable — counter advanced only). "
                f"Next #: {self._winner_next_spin}.\n"
            )
            self._push_undo_skip_counter_only(n)
            self._refresh_overlay_winner_session_list()

    def _browse(self) -> None:
        try:
            initial = self._list_path().parent
        except OSError:
            initial = script_dir()
        path = filedialog.askopenfilename(
            title="Step 1 — Choose your prize list file",
            initialdir=str(initial if initial.is_dir() else script_dir()),
            filetypes=[
                ("Prize lists", "*.xlsx *.txt *.tsv"),
                ("Excel workbook", "*.xlsx"),
                ("Text / TSV", "*.txt *.tsv"),
                ("All files", "*.*"),
            ],
        )
        if path:
            self.path_var.set(path)
            self._invalidate_session()
            self._refresh_setup_status()

    def _open_in_explorer(self, path: Path, *, what: str) -> None:
        target = path if path.is_dir() else path.parent
        if not target.is_dir():
            messagebox.showinfo(
                what,
                "That location does not exist yet.\n\n"
                "Use Choose… to pick a valid file or folder first.",
            )
            return
        try:
            os.startfile(str(target))
        except OSError as e:
            messagebox.showerror(what, f"Could not open folder:\n{e}")

    def _open_list_file_folder(self) -> None:
        try:
            p = self._list_path()
        except OSError:
            messagebox.showinfo("Prize list", "Choose a prize list file first.")
            return
        self._open_in_explorer(p, what="Prize list folder")

    def _open_images_folder(self) -> None:
        try:
            p = self._images_base_path()
        except OSError:
            messagebox.showinfo("Images folder", "Choose an images folder first.")
            return
        self._open_in_explorer(p, what="Images folder")

    def _setup_match_wheel_preset_from_file(self) -> None:
        """Pick a prize list file; use its name as the preset id and apply paths."""
        try:
            start = self._list_path().parent
        except OSError:
            start = script_dir()
        picked = filedialog.askopenfilename(
            title="Match wheel preset from prize list file",
            initialdir=str(start),
            filetypes=[
                ("Excel / prize list", "*.xlsx"),
                ("All files", "*.*"),
            ],
        )
        if not picked:
            return
        self._setup_apply_wheel_preset(list_file=Path(picked))

    def _setup_apply_wheel_preset(self, *, list_file: Path | None = None) -> None:
        previous_wid = draw_prize.normalize_wheel_id(
            getattr(self, "_active_wheel_preset_id", "") or ""
        )
        if list_file is not None:
            try:
                wid, list_path, images_dir = draw_prize.wheel_preset_paths_from_list_file(list_file)
            except draw_prize.PrizeDrawError as e:
                messagebox.showinfo("Wheel preset", str(e))
                return
            self.wheel_preset_var.set(wid)
        else:
            raw = self.wheel_preset_var.get()
            try:
                list_path, images_dir = draw_prize.wheel_preset_paths(raw, script_dir())
            except draw_prize.PrizeDrawError as e:
                messagebox.showinfo("Wheel preset", str(e))
                return
            wid = draw_prize.normalize_wheel_id(raw)
        self.path_var.set(str(list_path.resolve()))
        self.images_path_var.set(str(images_dir.resolve()))
        self._invalidate_session()
        self._on_images_path_write()
        self._refresh_setup_status()
        self._active_wheel_preset_id = wid
        if previous_wid != wid:
            self._create_and_attach_winner_session_for_preset(wid)
        wp = getattr(self, "_winner_log_path", None)
        if wp is not None and Path(wp).is_file():
            winner_bit = f"Winner log: {Path(wp).name}"
        else:
            winner_bit = "Winner log: none — install openpyxl or choose a spreadsheet"
        missing: list[str] = []
        if not list_path.is_file():
            missing.append(list_path.name)
        if not images_dir.is_dir():
            missing.append(f"{images_dir.name}/ folder")
        if missing:
            missing_where = (
                "Add them in the same folder as the prize list, then spin."
                if list_file is not None
                else "Add them next to the app, then spin."
            )
            self._log(
                f"Wheel preset {wid}: {list_path.name} · images {images_dir.name}/ · {winner_bit} · "
                f"not found yet: {', '.join(missing)} — {missing_where}\n"
            )
        else:
            self._log(
                f"Wheel preset {wid}: {list_path.name} + {images_dir.name}/ · {winner_bit}\n"
            )
        if self._setup_expanded.get():
            self.after_idle(self._overlay_resize_for_setup)
        self.after_idle(self._update_draw_buttons_for_supply)
        self._schedule_save_app_state()

    def _cancel_scheduled_app_state_save(self) -> None:
        aid = getattr(self, "_app_state_save_after", None)
        if aid is not None:
            try:
                self.after_cancel(aid)
            except tk.TclError:
                pass
            self._app_state_save_after = None

    def _setup_reset_persisted_settings(self) -> None:
        """Delete energy_break_state.json and clear in-app session fields (no path defaults)."""
        if self._busy or self._pending_super is not None:
            messagebox.showinfo(
                "Reset saved session",
                "Finish or cancel the current spin (including Super / Reroll / Keep) first.",
            )
            return
        if getattr(self, "_backfill_target_spot", None) is not None:
            messagebox.showinfo(
                "Reset saved session",
                'Cancel "Fill a skipped spot" first (Cancel fill mode in Spin & controls).',
            )
            return
        if not messagebox.askyesno(
            "Reset saved session",
            "Delete saved session data and clear this run?\n\n"
            f"Removes {app_local_state.STATE_FILENAME} (wheel preset, file paths, winner spreadsheet).\n"
            "Prize list and images paths are cleared (not reset to Input List / Images).\n"
            "Prize lists, wheel presets, and winner_sessions/ files on disk are not deleted.",
        ):
            return
        self._cancel_scheduled_app_state_save()
        if not app_local_state.delete_state(script_dir()):
            messagebox.showerror(
                "Reset saved session",
                f"Could not delete {app_local_state.STATE_FILENAME}.",
            )
            return
        self.wheel_preset_var.set("")
        self._active_wheel_preset_id = ""
        self.path_var.set("")
        self.images_path_var.set("")
        self._winner_log_path = None
        self._winner_next_spin = 1
        self._clear_undo_spin()
        self._update_spin_counter_label()
        self._refresh_winner_log_path_label()
        self._refresh_overlay_winner_session_list()
        self._invalidate_session()
        self._on_images_path_write()
        self._refresh_setup_status()
        self._update_draw_buttons_for_supply()
        if self._setup_expanded.get():
            self.after_idle(self._overlay_resize_for_setup)
        self._cancel_scheduled_app_state_save()
        app_local_state.delete_state(script_dir())
        self._log(
            f"Saved session cleared ({app_local_state.STATE_FILENAME} deleted). "
            "Apply a wheel preset or set prize list, images folder, and winner spreadsheet before spinning.\n"
        )
        messagebox.showinfo(
            "Reset saved session",
            "Session memory cleared. Set paths in Show setup or apply a wheel preset before spinning.",
        )

    def _refresh_setup_status(self) -> None:
        if not hasattr(self, "_setup_list_status"):
            return
        ok_fg = "#86efac"
        warn_fg = WHEEL_POINTER
        bad_fg = WHEEL_ACCENT

        try:
            p = self._list_path()
        except OSError:
            p = None
        if p is None or not p.is_file():
            list_raw = self.path_var.get().strip() if hasattr(self, "path_var") else ""
            if not list_raw:
                list_msg = "⚠ Prize list not set — apply a wheel preset or choose a prize list."
            else:
                list_msg = "⚠ Prize list not found — click Choose prize list… or apply a wheel preset."
            self._setup_list_status.configure(text=list_msg, fg=warn_fg)
        else:
            try:
                _, rows = draw_prize.load_rows(p)
                types_n = len(rows)
                units = sum(q for _, q, _ in rows)
                in_draw = sum(q for _, q, _ in rows if q > 0)
                self._setup_list_status.configure(
                    text=(
                        f"✓ Found: {p.name}  ·  {types_n} prize type(s)  ·  "
                        f"{units} total qty  ·  {in_draw} left to draw"
                    ),
                    fg=ok_fg,
                )
            except draw_prize.PrizeDrawError as e:
                self._setup_list_status.configure(
                    text=f"⚠ Could not read list: {e}",
                    fg=bad_fg,
                )
            except OSError as e:
                self._setup_list_status.configure(
                    text=f"⚠ Could not read list: {e}",
                    fg=bad_fg,
                )

        images_raw = self.images_path_var.get().strip() if hasattr(self, "images_path_var") else ""
        if not images_raw:
            self._setup_images_status.configure(
                text="⚠ Images folder not set — apply a wheel preset or choose an images folder.",
                fg=warn_fg,
            )
            return
        try:
            img_base = self._images_base_path()
        except OSError:
            self._setup_images_status.configure(
                text="⚠ Images folder not set — apply a wheel preset or choose an images folder.",
                fg=warn_fg,
            )
            return
        if not img_base.is_dir():
            self._setup_images_status.configure(
                text="⚠ Images folder not found — click Choose images folder…",
                fg=warn_fg,
            )
            return
        exts = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
        try:
            n_img = sum(
                1
                for f in img_base.iterdir()
                if f.is_file() and f.suffix.lower() in exts
            )
        except OSError:
            n_img = 0
        sample = ""
        try:
            names = sorted(
                f.name for f in img_base.iterdir() if f.is_file() and f.suffix.lower() in exts
            )[:3]
            if names:
                sample = "  e.g. " + ", ".join(names)
                if n_img > 3:
                    sample += ", …"
        except OSError:
            pass
        if n_img == 0:
            self._setup_images_status.configure(
                text=f"⚠ Folder exists but no .jpg/.png images found in {img_base.name}{sample}",
                fg=warn_fg,
            )
        else:
            self._setup_images_status.configure(
                text=f"✓ {n_img} image file(s) in {img_base.name}{sample}",
                fg=ok_fg,
            )

    def _images_base_path(self) -> Path:
        """Folder used to resolve relative img paths from the prize list (e.g. images/GrassEnergy.jpg)."""
        if hasattr(self, "images_path_var"):
            raw = self.images_path_var.get().strip()
            if not raw:
                raise OSError("Images folder path is not set")
            try:
                p = Path(raw).expanduser()
                if p.is_dir():
                    return p.resolve()
            except OSError:
                pass
            raise OSError("Images folder path is not set")
        fallback = script_dir() / "Images"
        return fallback.resolve() if fallback.is_dir() else script_dir().resolve()

    def _browse_images(self) -> None:
        try:
            initial = self._images_base_path()
        except OSError:
            initial = script_dir()
        path = filedialog.askdirectory(
            title="Step 2 — Choose the folder with prize pictures",
            initialdir=str(initial if initial.is_dir() else script_dir()),
            mustexist=True,
        )
        if path:
            self.images_path_var.set(path)
            self._on_images_path_write()

    def _on_images_path_write(self, *_args: object) -> None:
        self._wheel_image_cache.clear()
        self._inventory_image_cache.clear()
        self._maybe_refresh_idle_wheel()
        if getattr(self, "remaining_inner", None) is not None:
            try:
                if self.remaining_inner.winfo_exists():
                    self._refresh_remaining_skus_panel()
            except tk.TclError:
                pass
        self._refresh_setup_status()
        self._schedule_save_app_state()

    def _invalidate_session(self) -> None:
        self._session = None
        self._session_target = ""
        self._wheel_image_cache.clear()
        self._inventory_image_cache.clear()

    @staticmethod
    def _sku_to_img_map_from_rows(rows: list[tuple[str, int, str]]) -> dict[str, str]:
        out: dict[str, str] = {}
        for sku, _q, img in rows:
            if sku:
                out[sku] = img or ""
        return out

    def _resolve_prize_image_path(self, ref: str) -> Path | None:
        refn = _normalize_image_ref(ref)
        if not refn or refn.lower().startswith(("http://", "https://")):
            return None
        try:
            list_base = self._list_path().parent
            app_base = script_dir()
            images_base = self._images_base_path()
        except OSError:
            return None
        ref_path = Path(refn)
        rel_join = _path_for_project_join(refn)
        file_name = Path(rel_join).name
        candidates: list[Path] = []
        seen: set[str] = set()

        def add_candidate(p: Path) -> None:
            try:
                key = str(p.resolve())
            except OSError:
                key = str(p)
            if key not in seen:
                seen.add(key)
                candidates.append(p)

        if ref_path.is_absolute():
            add_candidate(ref_path.expanduser().resolve())
        add_candidate(images_base / rel_join)
        if file_name and file_name != rel_join:
            add_candidate(images_base / file_name)
        add_candidate((list_base / rel_join).resolve())
        add_candidate((app_base / rel_join).resolve())
        add_candidate((Path.cwd() / rel_join).resolve())
        if not ref_path.is_absolute():
            add_candidate(ref_path.expanduser().resolve())
        for cand in candidates:
            if cand.is_file():
                return cand
        return None

    def _wheel_thumb_max_px(self) -> tuple[int, int]:
        h = int(self.wheel_canvas.cget("height"))
        cell_h = max(32, h - WHEEL_STRIP_PAD_TOP - WHEEL_STRIP_PAD_BOTTOM - 4)
        cw = float(self.WHEEL_CELL)
        pl, pr = self._wheel_cell_xmargins()
        pad = float(pl + pr) + 8.0
        return max(28, int(cw - pad)), max(28, int(cell_h - pad))

    def _pil_image_to_tk_photo(self, im, mw: int, mh: int) -> object | None:
        try:
            from PIL import Image as PIM
            from PIL import ImageTk as PIMTk
        except ImportError:
            return None
        w, h = im.size
        scale = min(mw / max(w, 1), mh / max(h, 1), 1.0)
        nw, nh = max(1, int(w * scale)), max(1, int(h * scale))
        try:
            resample = PIM.Resampling.LANCZOS
        except AttributeError:
            resample = PIM.LANCZOS  # type: ignore[attr-defined]
        rgba = im.convert("RGBA")
        resized = rgba.resize((nw, nh), resample)
        try:
            return PIMTk.PhotoImage(resized, master=self)
        except tk.TclError:
            rgb = PIM.new("RGB", resized.size, (255, 255, 255))
            rgb.paste(resized, mask=resized.split()[3])
            return PIMTk.PhotoImage(rgb, master=self)

    def _pil_image_to_wheel_photo(self, im) -> object | None:
        mw, mh = self._wheel_thumb_max_px()
        return self._pil_image_to_tk_photo(im, mw, mh)

    def _open_pil_image_from_ref(self, refn: str):
        """Return a PIL Image from URL or disk path, or None."""
        if not refn or not _pillow_runtime_available():
            return None
        try:
            from PIL import Image as PILImage
        except ImportError:
            return None
        if refn.lower().startswith(("http://", "https://")):
            try:
                req = urllib.request.Request(
                    refn,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
                    },
                )
                ctx = ssl.create_default_context()
                with urllib.request.urlopen(req, timeout=20, context=ctx) as resp:
                    data = resp.read()
            except Exception:
                return None
            if len(data) < 32 or _looks_like_html_head(data):
                return None
            try:
                im = PILImage.open(BytesIO(data))
                im.load()
                return im
            except Exception:
                return None
        path = self._resolve_prize_image_path(refn)
        if path is None:
            return None
        try:
            im = PILImage.open(path)
            im.load()
            return im
        except Exception:
            return None

    def _wheel_photo_for_normalized_ref(self, refn: str) -> object | None:
        """Load (or return cached) ImageTk.PhotoImage for wheel cell; caches by normalized ref."""
        if not refn or not _pillow_runtime_available():
            return None
        if refn in self._wheel_image_cache:
            return self._wheel_image_cache[refn]
        im = self._open_pil_image_from_ref(refn)
        if im is None:
            return None
        photo = self._pil_image_to_wheel_photo(im)
        if photo is not None:
            self._wheel_image_cache[refn] = photo
        return photo

    def _inventory_photo_for_normalized_ref(self, refn: str, thumb_px: int) -> object | None:
        """Square thumb for prize-board grid; cache key includes thumb size."""
        if not refn or not _pillow_runtime_available():
            return None
        key = (refn, thumb_px)
        if key in self._inventory_image_cache:
            return self._inventory_image_cache[key]
        im = self._open_pil_image_from_ref(refn)
        if im is None:
            return None
        photo = self._pil_image_to_tk_photo(im, thumb_px, thumb_px)
        if photo is not None:
            self._inventory_image_cache[key] = photo
        return photo

    def _ensure_wheel_photos_for_strip(self, strip: list[str]) -> None:
        smap = self._wheel_sku_to_img
        done: set[str] = set()
        for sku in strip:
            if sku in ("—", "?", "···", ""):
                continue
            ref = (smap or {}).get(sku, "")
            refn = _normalize_image_ref(ref)
            if not refn or refn in done or refn in self._wheel_image_cache:
                if refn:
                    done.add(refn)
                continue
            done.add(refn)
            self._wheel_photo_for_normalized_ref(refn)

    def _ensure_session(self) -> draw_prize.FileDrawSession:
        key = str(self._list_path())
        if self._session is not None and self._session_target == key:
            return self._session
        self._session = draw_prize.open_draw_session(Path(key))
        self._session_target = key
        return self._session

    def _clear_animation(self) -> None:
        if self._anim_after is not None:
            try:
                self.after_cancel(self._anim_after)
            except tk.TclError:
                pass
            self._anim_after = None

    def _set_busy(self, busy: bool) -> None:
        self._busy = busy
        if busy:
            self._cancel_wheel_idle_drift()
        entry_state = tk.DISABLED if busy else tk.NORMAL
        for ent_name in ("path_entry", "images_path_entry"):
            ent = getattr(self, ent_name, None)
            if ent is not None:
                try:
                    ent.configure(state=entry_state)
                except tk.TclError:
                    pass
        if busy:
            if hasattr(self, "spin_btn"):
                self._btn_set_busy(self.spin_btn, True)
                self._btn_set_busy(self.super_btn, True)
                if hasattr(self, "reroll_btn") and self.reroll_btn.winfo_ismapped():
                    self._btn_set_busy(self.reroll_btn, True)
                    self._btn_set_busy(self.keep_btn, True)
                self._btn_set_busy(self.skip_spot_btn, True)
            if hasattr(self, "undo_spin_btn"):
                self._btn_set_busy(self.undo_spin_btn, True)
        elif self._pending_super is not None:
            self._style_super_controls_active()
        else:
            self._update_draw_buttons_for_supply()
        self._sync_undo_spin_button()

    def _hide_super_panel(self) -> None:
        self._pending_super = None
        self._super_reroll_used = False
        if not self._busy:
            self._update_draw_buttons_for_supply()

    def _show_super_panel(self, session: draw_prize.FileDrawSession, result: draw_prize.SpinResult) -> None:
        self._pending_super = (session, result)
        self._style_super_controls_active()

    def _worker_pick(
        self,
        on_ok: object,
        on_err: object,
    ) -> None:
        def run() -> None:
            try:
                session = self._ensure_session()
                session.refresh()
                if not any(q > 0 for _, q, _ in session.rows):
                    raise draw_prize.PrizeDrawError("No prizes left (all quantities are 0).")
                i = draw_prize.pick_sku(session.rows)
                sku, qty, img = session.rows[i]
                r = draw_prize.SpinResult(i, sku, qty, img)
                self.after(0, lambda: on_ok(session, r, session.rows))
            except draw_prize.PrizeDrawError as e:
                self.after(0, lambda: on_err(str(e)))
            except Exception as e:
                self.after(0, lambda: on_err(f"{type(e).__name__}: {e}"))

        threading.Thread(target=run, daemon=True).start()

    def _worker_commit(
        self,
        session: draw_prize.FileDrawSession,
        result: draw_prize.SpinResult,
        on_done: object,
    ) -> None:
        def run() -> None:
            try:
                if not self.dry_run_var.get():
                    session.commit(result)
                self.after(0, lambda: on_done(None))
            except draw_prize.PrizeDrawError as e:
                self.after(0, lambda: on_done(str(e)))
            except Exception as e:
                self.after(0, lambda: on_done(f"{type(e).__name__}: {e}"))

        threading.Thread(target=run, daemon=True).start()

    def _on_spin(self) -> None:
        if self._busy:
            return
        block = self._winner_session_spin_block_reason()
        if block:
            messagebox.showinfo("Cannot spin", block)
            return
        self._hide_super_panel()
        self._set_busy(True)
        self._cancel_pulse()
        self._set_wheel_status("Loading picks from your list…", WHEEL_TITLE)
        self._wheel_loading_preview_setup()

        def ok(
            session: draw_prize.FileDrawSession,
            r: draw_prize.SpinResult,
            rows: list[tuple[str, int, str]],
        ) -> None:
            def after_anim() -> None:
                dry = self.dry_run_var.get()
                self._log(f"Prize: {r.sku}  (was Qty {r.qty}){'  [dry run]' if dry else ''}\n")

                def commit_done(err: str | None) -> None:
                    self._set_busy(False)
                    if err:
                        messagebox.showerror("Could not save", err)
                        self._log(f"Error: {err}\n")
                        self._invalidate_session()
                        self._wheel_show_error(str(err))
                        self._refresh_prizes_label()
                    else:
                        if not dry:
                            self._log(f"Updated: {r.sku} Qty -> {max(0, r.qty - 1)}\n")
                            bf = getattr(self, "_backfill_target_spot", None)
                            if bf is not None:
                                winner_ok = self._record_spin_winner(r.sku)
                                self._push_undo_backfill(session.path.resolve(), r, int(bf), winner_ok)
                                if winner_ok:
                                    self._set_obs_wheel_result(int(bf), r.sku)
                                    self.cancel_fill_skipped_spot_mode(silent=True)
                                    self._log(
                                        f"Filled skipped Spot #{bf} in the winner session with {r.sku}.\n"
                                    )
                                else:
                                    self._log(
                                        "Prize list was saved, but the winner sheet row could not be filled "
                                        "(try closing Excel). Fill-skip mode is still armed for another try.\n"
                                    )
                            else:
                                spot_before = self._winner_next_spin
                                self._set_obs_wheel_result(spot_before, r.sku)
                                winner_ok = self._record_spin_winner(r.sku)
                                self._push_undo_spin(session.path.resolve(), r, spot_before, winner_ok)
                            self._refresh_prizes_label()
                            last = ""
                            uq = self._winnable_units_total()
                            if uq is not None and uq == 0:
                                last = "  ·  That was the last prize — list is empty."
                            self._set_wheel_status(f"Winner: {r.sku}", WHEEL_POINTER)
                            self._pulse_wheel_win()
                        else:
                            self._set_wheel_status(
                                f"{r.sku}  ·  dry run (file unchanged), was Qty {r.qty}",
                                WHEEL_TITLE,
                            )

                if dry:
                    self._log("Dry run: file not updated.\n")
                    commit_done(None)
                else:
                    self._worker_commit(session, r, commit_done)

            self._run_wheel_spin(
                rows,
                r.sku,
                f"Landed on: {r.sku}  ·  Qty {r.qty}" + ("  ·  DRY RUN" if self.dry_run_var.get() else ""),
                after_anim,
                self._sku_to_img_map_from_rows(session.rows),
            )

        def err(msg: str) -> None:
            self._set_busy(False)
            self._invalidate_session()
            self._refresh_prizes_label()
            units = self._winnable_units_total()
            if units is not None and units == 0:
                messagebox.showinfo(
                    "Prizes finished",
                    "Every prize has been given out. Add quantities to your list file to draw again.",
                )
            else:
                self._wheel_show_error(str(msg))
                messagebox.showerror("Spin", msg)
            self._log(f"Error: {msg}\n")

        self._worker_pick(ok, err)

    def _on_super_spin(self) -> None:
        if self._busy:
            return
        block = self._winner_session_spin_block_reason()
        if block:
            messagebox.showinfo("Cannot super spin", block)
            return
        if getattr(self, "_backfill_target_spot", None) is not None:
            messagebox.showinfo(
                "Super spin",
                "Fill-skip mode uses a normal SPIN only. Cancel fill mode first, or complete that spin.",
            )
            return
        self._hide_super_panel()
        self._set_busy(True)
        self._cancel_pulse()
        self._set_wheel_status("Super spin — loading picks…", WHEEL_TITLE)
        self._wheel_loading_preview_setup()

        def ok(
            session: draw_prize.FileDrawSession,
            r: draw_prize.SpinResult,
            rows: list[tuple[str, int, str]],
        ) -> None:
            def after_anim() -> None:
                self._show_super_panel(session, r)
                self._set_busy(False)
                self._log(f"Super spin result: {r.sku}  (was Qty {r.qty})\n")
                self._set_wheel_status("Keep or Reroll!", WHEEL_FG)

            self._run_wheel_spin(
                rows,
                r.sku,
                f"Showing: {r.sku}  ·  Qty {r.qty}",
                after_anim,
                self._sku_to_img_map_from_rows(session.rows),
            )

        def err(msg: str) -> None:
            self._set_busy(False)
            self._invalidate_session()
            self._refresh_prizes_label()
            units = self._winnable_units_total()
            if units is not None and units == 0:
                messagebox.showinfo(
                    "Prizes finished",
                    "Every prize has been given out. Add quantities to your list file to draw again.",
                )
            else:
                self._wheel_show_error(str(msg))
                messagebox.showerror("Super spin", msg)
            self._log(f"Error: {msg}\n")

        self._worker_pick(ok, err)

    def _on_reroll(self) -> None:
        if self._busy or self._pending_super is None or self._super_reroll_used:
            return
        self._set_busy(True)
        self._cancel_pulse()
        self._set_wheel_status("Rerolling — loading picks…", WHEEL_TITLE)
        self._wheel_loading_preview_setup()

        def ok(
            s: draw_prize.FileDrawSession,
            r: draw_prize.SpinResult,
            rows: list[tuple[str, int, str]],
        ) -> None:
            def after_anim() -> None:
                self._super_reroll_used = True
                self._show_super_panel(s, r)
                self._set_busy(False)
                self._log(f"Reroll: {r.sku}  (was Qty {r.qty})\n")
                self._set_wheel_status(
                    f"No more rerolls — KEEP to save −1. Pointer: {r.sku}  ·  qty {r.qty}",
                    WHEEL_FG,
                )

            self._run_wheel_spin(
                rows,
                r.sku,
                f"Reroll: {r.sku}  ·  Qty {r.qty}",
                after_anim,
                self._sku_to_img_map_from_rows(s.rows),
            )

        def err(msg: str) -> None:
            self._set_busy(False)
            self._invalidate_session()
            self._reset_wheel_idle()
            self._refresh_prizes_label()
            units = self._winnable_units_total()
            if units is not None and units == 0:
                messagebox.showinfo(
                    "Prizes finished",
                    "Every prize has been given out. Add quantities to your list file to draw again.",
                )
            else:
                self._wheel_show_error(str(msg))
                messagebox.showerror("Reroll", msg)
            self._log(f"Error: {msg}\n")

        self._worker_pick(ok, err)

    def _on_keep(self) -> None:
        if self._busy or self._pending_super is None:
            return
        session, result = self._pending_super
        self._set_busy(True)
        dry = self.dry_run_var.get()
        self._log(f"Keep: {result.sku}  (was Qty {result.qty}){'  [dry run]' if dry else ''}\n")

        def commit_done(err: str | None) -> None:
            self._hide_super_panel()
            self._set_busy(False)
            if err:
                messagebox.showerror("Could not save", err)
                self._log(f"Error: {err}\n")
                self._invalidate_session()
                self._wheel_show_error(str(err))
                self._refresh_prizes_label()
            else:
                if not dry:
                    self._log(f"Updated: {result.sku} Qty -> {max(0, result.qty - 1)}\n")
                    spot_before = self._winner_next_spin
                    self._set_obs_wheel_result(spot_before, result.sku)
                    winner_ok = self._record_spin_winner(result.sku)
                    self._push_undo_spin(session.path.resolve(), result, spot_before, winner_ok)
                    self._refresh_prizes_label()
                    last = ""
                    uq = self._winnable_units_total()
                    if uq is not None and uq == 0:
                        last = "  ·  That was the last prize — list is empty."
                    self._set_wheel_status(
                        f"Kept: {result.sku}  ·  was {result.qty}  →  now {max(0, result.qty - 1)}{last}",
                        WHEEL_POINTER,
                    )
                    self._pulse_wheel_win()
                else:
                    self._log("Dry run: file not updated.\n")
                    self._set_wheel_status(
                        f"{result.sku}  ·  dry run — qty still {result.qty} on file",
                        WHEEL_TITLE,
                    )

        if dry:
            commit_done(None)
        else:
            self._worker_commit(session, result, commit_done)


def _close_fill_skipped_dialog(top: tk.Toplevel) -> None:
    try:
        top.grab_release()
    except tk.TclError:
        pass
    try:
        top.destroy()
    except tk.TclError:
        pass


_SINGLE_INSTANCE_LOCK: object | None = None
_SINGLE_INSTANCE_MUTEX_NAME = "Global\\EnergyBreakSystem.SingleInstance.v1"
_ALREADY_RUNNING_TITLE = "Energy Break"
_ALREADY_RUNNING_MSG = (
    "Energy Break is already running.\n\n"
    "Close the other window first. If you do not see it, check Task Manager for "
    "python.exe or Energy Break, then try again.\n\n"
    "Only one copy can run so the HTML/OBS wheel port (8765) stays in sync."
)


def _focus_existing_energy_break_window() -> None:
    """Best-effort: bring the first instance to the foreground (Windows)."""
    if sys.platform != "win32":
        return
    try:
        import ctypes

        user32 = ctypes.windll.user32
        for title in ("Energy Break — Overlay", "Energy Break — Spin & controls"):
            hwnd = user32.FindWindowW(None, title)
            if hwnd:
                user32.ShowWindow(hwnd, 9)  # SW_RESTORE
                user32.SetForegroundWindow(hwnd)
                return
    except Exception:
        pass


def _show_already_running_message() -> None:
    if sys.platform == "win32":
        try:
            import ctypes

            ctypes.windll.user32.MessageBoxW(
                None,
                _ALREADY_RUNNING_MSG,
                _ALREADY_RUNNING_TITLE,
                0x40,
            )
            return
        except Exception:
            pass
    root = tk.Tk()
    root.withdraw()
    try:
        messagebox.showinfo(_ALREADY_RUNNING_TITLE, _ALREADY_RUNNING_MSG, parent=root)
    finally:
        root.destroy()


def _acquire_single_instance_lock() -> bool:
    """
    Return True if this process is the only Energy Break instance.

    Uses a Windows named mutex when available; otherwise a non-blocking lock file.
    """
    global _SINGLE_INSTANCE_LOCK
    if _SINGLE_INSTANCE_LOCK is not None:
        return True

    if sys.platform == "win32":
        try:
            import ctypes

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
            kernel32.CreateMutexW.argtypes = [
                ctypes.c_void_p,
                ctypes.c_bool,
                ctypes.c_wchar_p,
            ]
            kernel32.CreateMutexW.restype = ctypes.c_void_p
            kernel32.CloseHandle.argtypes = [ctypes.c_void_p]
            kernel32.CloseHandle.restype = ctypes.c_bool

            handle = kernel32.CreateMutexW(None, True, _SINGLE_INSTANCE_MUTEX_NAME)
            if not handle:
                return True
            if ctypes.get_last_error() == 183:  # ERROR_ALREADY_EXISTS
                kernel32.CloseHandle(handle)
                return False
            _SINGLE_INSTANCE_LOCK = handle
            return True
        except Exception:
            pass

    import tempfile

    lock_path = Path(tempfile.gettempdir()) / "energy-break-system.lock"
    try:
        fh = open(lock_path, "a+b")
    except OSError:
        return True
    try:
        if sys.platform == "win32":
            import msvcrt

            fh.seek(0)
            try:
                msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
            except OSError:
                fh.close()
                return False
        else:
            import fcntl

            try:
                fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except OSError:
                fh.close()
                return False
        fh.seek(0)
        fh.truncate()
        fh.write(str(os.getpid()).encode("ascii"))
        fh.flush()
        _SINGLE_INSTANCE_LOCK = fh
        return True
    except Exception:
        try:
            fh.close()
        except OSError:
            pass
        return True


def _ensure_single_instance() -> None:
    if _acquire_single_instance_lock():
        return
    _focus_existing_energy_break_window()
    _show_already_running_message()
    raise SystemExit(0)


def main() -> None:
    _ensure_single_instance()
    app = DrawPrizeApp()
    app.mainloop()


if __name__ == "__main__":
    main()
